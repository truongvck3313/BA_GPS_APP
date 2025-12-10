import time
import openpyxl
import os
from selenium.common.exceptions import NoSuchElementException
import subprocess
import var_app
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import logging
import module_gps_appv3
import json
import requests
from requests.auth import HTTPBasicAuth







def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename="bagps_app.log",
                                                     encoding='utf-8', mode='w')],  #w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)





def back(self):
    import login_app
    import minitor_app
    import route
    import vehicle
    import homepage_app
    import image_video_app
    import report_app
    import utilities

    var_app.driver.implicitly_wait(0.1)
    login_app.login.login_back(self)
    minitor_app.minitor_back()
    route.route_back()
    vehicle.vehicle_back()
    homepage_app.toolbar_back()
    image_video_app.image_back()
    report_app.general.report_back(self)
    utilities.general.utilities_back(self)



# @retry(tries=3, delay=2, backoff=1, jitter=5, )
def close_app():
    var_app.driver.implicitly_wait(0.1)
    try:
        var_app.driver.find_element(By.XPATH, var_app.allow).click()
    except:
        pass

    try:
        var_app.driver.find_element(By.XPATH, var_app.closeapp).click()
    except:
        pass
    try:
        var_app.driver.find_element(By.XPATH, var_app.wait).click()
    except:
        pass
    try:
        var_app.driver.find_element(By.XPATH, var_app.clearall).click()
    except:
        pass


    reset_driver()
    # try:
    #     var_app.driver.implicitly_wait(0.5)
    # except:
    #     var_app.restart_driver()
    #
    # time.sleep(1)
    # try:
    #     var_app.driver.find_element(By.XPATH, var_app.close_app).click()
    #     time.sleep(1)
    # except:
    #     pass
    #
    # try:
    #     var_app.driver.find_element(By.XPATH, var_app.ALLOW).click()
    #     time.sleep(1)
    # except:
    #     pass
    #
    #
    # try:
    #     wait = WebDriverWait(var_app.driver, 10)
    #     element = wait.until(EC.element_to_be_clickable((By.XPATH, var_app.bagps)))
    # except:
    #     reset_app_image()
    #
    # print("Đã đóng app")



def open_bagps():
    import subprocess
    import time

    adb_path = "adb"
    udid = "YOUR_DEVICE_ID"  # hoặc lấy từ var_app.udid nếu bạn có sẵn

    app_package = "vn.bagps.gpsmobile"
    app_activity = "crc64e73994338fce12ed.MainActivity"

    # Mở lại app
    subprocess.call([adb_path, "-s", udid, "shell", "am", "start", "-n", f"{app_package}/{app_activity}"])
    time.sleep(1)

    print("✅ App đã mở lại.")





def check_devices():
    adb_path = r"{}adb.exe".format(var_app.ldp)
    print(adb_path)
    # adb_path = r"D:\LDPlayer\LDPlayer9\adb.exe"
    # Liệt kê các thiết bị kết nối
    output = subprocess.check_output([adb_path, "devices"], text=True)
    print(output)

def reset_app_image():
    pass
    # # Click vào dấu x
    # timeout = 10  # thời gian tối đa để chờ (giây)
    # start_time = time.time()
    # iconx_button = None
    # reset_button1 = None
    # reset_button2 = None
    #
    # while time.time() - start_time < timeout:
    #     iconx_button = pyautogui.locateOnScreen(var_app.uploadpath + 'iconx_button.png', confidence=0.8)
    #     if iconx_button:
    #         pyautogui.click(iconx_button)
    #         print("Đã click nút X.")
    #         time.sleep(2)
    #         break
    #     time.sleep(1)  # nghỉ 1 giây rồi kiểm tra lại
    # if not iconx_button:
    #     print("Không tìm thấy nút X sau 10 giây.")
    #
    #
    # try:
    #     while time.time() - start_time < timeout:
    #         reset_button1 = pyautogui.locateOnScreen(var_app.uploadpath + 'reset_button1.png', confidence=0.8)
    #         if reset_button1:
    #             pyautogui.click(reset_button1)
    #             print("Đã click nút reset1.")
    #             time.sleep(25)
    #             break
    #     if not reset_button1:
    #         print("Không tìm thấy nút reset1a")
    # except:
    #     print("Không tìm thấy nút reset1b")
    #
    #
    # try:
    #     while time.time() - start_time < timeout:
    #         reset_button2 = pyautogui.locateOnScreen(var_app.uploadpath + 'reset_button2.png', confidence=0.8)
    #         if reset_button2:
    #             pyautogui.click(reset_button2)
    #             print("Đã click nút reset2.")
    #             time.sleep(25)
    #             break
    #     if not reset_button2:
    #         print("Không tìm thấy nút reset2a")
    # except:
    #     print("Không tìm thấy nút reset2b")
    # var_app.restart_driver()
    # try:
    #     var_app.driver.implicitly_wait(2)
    # except:
    #     var_app.restart_driver()






def swipe_scaled(driver, start_x_old, start_y_old, end_x_old, end_y_old, duration=500, old_width=900, old_height=1600):
    """
    Swipe từ tọa độ trên màn hình cũ sang màn hình hiện tại tự động scale.

    driver: Appium driver
    start_x_old, start_y_old: tọa độ bắt đầu trên màn hình cũ
    end_x_old, end_y_old: tọa độ kết thúc trên màn hình cũ
    duration: thời gian swipe (ms)
    old_width, old_height: kích thước màn hình cũ
    """
    size = driver.get_window_size()
    width_new = size['width']
    height_new = size['height']

    # Tính tỉ lệ tọa độ
    start_x_new = int(start_x_old / old_width * width_new)
    start_y_new = int(start_y_old / old_height * height_new)
    end_x_new = int(end_x_old / old_width * width_new)
    end_y_new = int(end_y_old / old_height * height_new)

    # Thực hiện swipe
    driver.swipe(start_x_new, start_y_new, end_x_new, end_y_new, duration)




from appium.webdriver.common.touch_action import TouchAction
def drag_scaled_steps(driver,
                      start_x_old, start_y_old,
                      end_x_old, end_y_old,
                      steps=4,
                      step_wait=300,
                      old_width=900, old_height=1600):

    size = driver.get_window_size()
    new_w, new_h = size['width'], size['height']

    # Scale toạ độ
    start_x = int(start_x_old / old_width * new_w)
    start_y = int(start_y_old / old_height * new_h)
    end_x   = int(end_x_old   / old_width * new_w)
    end_y   = int(end_y_old   / old_height * new_h)

    # Tính khoảng cách mỗi bước
    delta_x = (end_x - start_x) / steps
    delta_y = (end_y - start_y) / steps

    action = TouchAction(driver)

    # Nhấn giữ tại điểm đầu
    action.long_press(x=start_x, y=start_y).wait(step_wait)

    # Kéo từng bước (loop)
    for i in range(steps):
        new_x = int(start_x + delta_x * (i + 1))
        new_y = int(start_y + delta_y * (i + 1))
        action.move_to(x=new_x, y=new_y).wait(step_wait)

    # Thả ra
    action.release().perform()


def swipe_percent(driver, start_x_pct, start_y_pct, end_x_pct, end_y_pct, duration=500):
    size = driver.get_window_size()
    width = size['width']
    height = size['height']

    start_x = int(width * start_x_pct)
    start_y = int(height * start_y_pct)
    end_x = int(width * end_x_pct)
    end_y = int(height * end_y_pct)

    driver.swipe(start_x, start_y, end_x, end_y, duration)
#module_other_app.swipe_percent(var_app.driver, startX, startY, endX, endY, duration)




def tap_scaled(driver, coords_old, old_width=900, old_height=1600):
    """
    Tap vào tọa độ trên màn hình cũ, tự động scale sang màn hình hiện tại.

    driver: Appium driver
    coords_old: list [(x, y)] giống cú pháp tap cũ
    old_width, old_height: kích thước màn hình cũ
    """
    size = driver.get_window_size()
    print(f"size: {size}")
    width_new = size['width']
    height_new = size['height']

    coords_new = []
    for x_old, y_old in coords_old:
        x_new = int(x_old / old_width * width_new)
        y_new = int(y_old / old_height * height_new)
        print(f"x_new: {x_new}\ny_new: {y_new}")
        coords_new.append((x_new, y_new))
    driver.tap(coords_new)







def reset_driver():
    adb_path = "adb"

    main_app = {
        "vn.bagps.gpsmobile": "crc64e73994338fce12ed.MainActivity"
    }

    # Lấy thiết bị đầu tiên
    try:
        output = subprocess.check_output([adb_path, "devices"], text=True)
        devices = [line.split()[0] for line in output.strip().splitlines()[1:] if line.endswith("device")]
    except subprocess.CalledProcessError:
        print("⚠ Không thể chạy adb devices.")
        return

    if not devices:
        print("⚠ Không tìm thấy thiết bị.")
        return

    udid = devices[0]
    print(f"✅ Đang sử dụng thiết bị: {udid}")

    def adb_shell(cmd):
        return subprocess.check_output([adb_path, "-s", udid, "shell"] + cmd.split(), text=True)

    # Force-stop và khởi động lại app chính
    for pkg, activity in main_app.items():
        print(f"➤ Reset app chính: {pkg}")
        try:
            adb_shell(f"am force-stop {pkg}")  # Dừng app chính
            time.sleep(0.5)
            adb_shell(f"am start -n {pkg}/{activity}")  # Khởi động lại
            time.sleep(1)
            print(f"✅ App {pkg} đã được reset.")
        except subprocess.CalledProcessError as e:
            print(f"⚠ Không thể reset app {pkg}: {e}")

    print("✅ Reset xong app chính.")



    #reret adb
    try:
        var_app.driver.find_element(By.XPATH, var_app.login_user1)
    except Exception as e:
        logging.warning("⚠️ UiAutomator2 crash → Restart driver")
        logging.warning("e")
        var_app.restart_driver()



def fake_ip():
    from faker import Faker
    from faker.providers import internet
    fake = Faker()
    fake.add_provider(internet)
    print(fake.ipv4_private())





def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_app.timerun)
        writeData1(var_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        if var_app.timerun == "":
            writeData1(var_app.path_luutamthoi, "Sheet2", 2, 2, timerun)
        else:
            writeData1(var_app.path_luutamthoi, "Sheet2", 2, 2, var_app.timerun)


        if var_app.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_app.timerun == "":
            break





def clearData(file,sheetName, data, ketqua, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1300):
        i += 1
        i = str(i)
        sheet["F"+i] = data
        sheet["G"+i] = ketqua
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)


def clearData_luutamthoi(file,sheetName, overview, detail, api):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 1
    while (i < 300):
        i += 1
        i = str(i)
        sheet["B"+i] = overview
        sheet["C"+i] = detail
        sheet["D"+i] = api
        i = int(i)
    wordbook.save(file)




def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value




def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)




def writeData1(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)




def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_app.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)

    module_gps_appv3.check_casenone()
    module_gps_appv3.check_casefail()
    module_gps_appv3.check_casepass()

    mucnghiemtrong = str(readData(var_app.path_luutamthoi, 'Sheet1', 85, 2))
    tong_case_trong = str(readData(var_app.path_luutamthoi, 'Sheet1', 86, 2))
    case_fail = str(readData(var_app.path_luutamthoi, 'Sheet1', 167, 2))
    case_pass = str(readData(var_app.path_luutamthoi, 'Sheet1', 177, 2))

    print("Pass: "+ case_pass)
    print("Fail: "+ case_fail)
    print("Trống: "+ tong_case_trong)
    print("Nghiêm trọng: "+ mucnghiemtrong)




    thoigianbatdauchay = str(readData(var_app.path_luutamthoi, "Sheet2", 2, 2))

    print(thoigianbatdauchay)
    # if case_fail >= 1:
    driver2.get("https://web.telegram.org/a/#-4248738484")
    time.sleep(2)
    case_pass = str(case_pass)
    case_fail = str(case_fail)
    driver2.ele(var_app.hopthoai).click()
    time.sleep(0.5)
    time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    driver2.ele(var_app.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- ModeTest: " + var_app.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)
    driver2.ele(var_app.hopthoai_input).input(Keys.ENTER)
    time.sleep(1)
    driver2.ele(var_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_app.hopthoai_iconlink_file).click()
    time.sleep(1)
    subprocess.Popen(var_app.uploadpath+"GPS_APPChecklistForAutoTest.exe")
    time.sleep(2)
    driver2.ele(var_app.hopthoai_send).click()
    time.sleep(2)
    driver2.ele(var_app.hopthoai_iconlink).click()
    time.sleep(1)
    driver2.ele(var_app.hopthoai_iconlink_file).click()
    time.sleep(2)
    subprocess.Popen(var_app.uploadpath+"bagps_app.exe")
    time.sleep(2)
    driver2.ele(var_app.hopthoai_send).click()
    time.sleep(1)





def get_version():
    version = str(readData(var_app.path_luutamthoi, 'Sheet1', 115, 2))
    print(f"version old: {version}")

    if version == "None" or version is None:
        workbook = openpyxl.load_workbook(var_app.checklistpath)
        sheet = workbook["Checklist"]
        rownum = 9
        while rownum < 3000:
            rownum += 1
            if sheet[f"A{rownum}"].value == "Account01":
                version = sheet[f"F{rownum}"].value
                print("version mới từ checklist:", version)  # ← THÊM DÒNG NÀY
                break  # ← QUAN TRỌNG: THOÁT WHILE, KHÔNG RETURN
    return version  # ← ĐƯA RETURN RA CUỐI HÀM



def viber_send_text():
    module_gps_appv3.check_casenone()
    module_gps_appv3.check_casefail()
    module_gps_appv3.check_casepass()

    mucnghiemtrong = str(readData(var_app.path_luutamthoi, 'Sheet1', 85, 2))
    tong_case_trong = str(readData(var_app.path_luutamthoi, 'Sheet1', 86, 2))
    case_fail = str(readData(var_app.path_luutamthoi, 'Sheet1', 167, 2))
    case_pass = str(readData(var_app.path_luutamthoi, 'Sheet1', 177, 2))
    version = str(readData(var_app.path_luutamthoi, 'Sheet1', 115, 2))
    if version == "None" or version is None:
        version = get_version()

    print("version: "+ version)
    print("Pass: "+ case_pass)
    print("Fail: "+ case_fail)
    print("Trống: "+ tong_case_trong)
    print("Nghiêm trọng: "+ mucnghiemtrong)

    thoigianbatdauchay = str(readData(var_app.path_luutamthoi, "Sheet2", 2, 2))
    time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)

    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- ModeTest: " + var_app.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)

    AUTH_TOKEN = "54c4302d943110cf-585274afa3260c1c-fc90f7e41b1c03d4"  # id Cảnh báo Autotest App gps Customer
    FROM_USER_ID = "HAtk2LeVK9qaafCeG+PRDA=="


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    print(AUTH_TOKEN)
    print(FROM_USER_ID)


    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return


    # 2. Gửi tin nhắn văn bản
    send_url = "https://chatapi.viber.com/pa/post"
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,
        "type": "text",
        "text": ("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              f"\n- Version BA GPS: {version}"
                                              "\n- ModeTest: " + var_app.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case Trống: "+ tong_case_trong+
                                              "\n- Số case False Nghiêm trọng: "+ mucnghiemtrong)}

    headers = {
        "Content-Type": "application/json"}

    response = requests.post(send_url, json=payload, headers=headers)

    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())


def check_user_id():
    res = requests.post(
        "https://chatapi.viber.com/pa/get_account_info",
        headers={"X-Viber-Auth-Token": "54c4302d943110cf-585274afa3260c1c-fc90f7e41b1c03d4"}
    )
    print(res.json())


# def upload_to_catbox(file_path):
#     url = "https://catbox.moe/user/api.php"
#     files = {
#         'fileToUpload': open(file_path, 'rb')
#     }
#     data = {
#         'reqtype': 'fileupload'
#     }
#
#     try:
#         response = requests.post(url, files=files, data=data)
#         response.raise_for_status()
#     except requests.RequestException as e:
#         print("❌ Lỗi upload:", e)
#         return None
#
#     direct_link = response.text.strip()
#     # Catbox trả về URL trực tiếp của file, vd: https://files.catbox.moe/abc123.png
#     print(f"✅ Upload thành công! Link tải trực tiếp:\n{direct_link}")
#     return direct_link


def upload_pixeldrain_auth(filepath):
    API_KEY = "c567bb13-f4c0-4aac-b9bd-c8add1e467fc"  # Thay bằng key thật

    with open(filepath, "rb") as f:
        res = requests.post(
            "https://pixeldrain.com/api/file",
            files={"file": f},
            auth=HTTPBasicAuth('', API_KEY)
        )
        res_json = json.loads(res.text)
        file_id = res_json["id"]
        link_download = (f"https://pixeldrain.com/api/file/{file_id}")
        print(link_download)
        return link_download



def send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, file_path):
    file_url = upload_pixeldrain_auth(file_path)
    if not file_url:
        print("⚠️ Không thể upload file. Hủy gửi.")
        return

    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,           # Viber user ID người nhận
        "min_api_version": 1,
        "tracking_data": "",               # Có thể để chuỗi rỗng nếu không dùng tracking
        "type": "file",
        "media": file_url,
        "size": file_size,
        "file_name": file_name}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())


def viber_send_file():
    # ==== Ví dụ sử dụng ====
    AUTH_TOKEN = "54c4302d943110cf-585274afa3260c1c-fc90f7e41b1c03d4"  # id Cảnh báo Autotest App gps
    FROM_USER_ID = "HAtk2LeVK9qaafCeG+PRDA=="


    FILE_PATH_checklisst = var_app.checklistpath  # Thay bằng đường dẫn file thật
    FILE_PATH_log = var_app.logpath  # Thay bằng đường dẫn file thật

    print(var_app.checklistpath)
    print(var_app.logpath)


    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_checklisst)
    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_log)


def send_viber():
    viber_send_text()
    viber_send_file()






















def delete_image():
    path = os.path.join(var_app.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def get_datachecklist(ma):
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 3000):
        rownum += 1
        rownum = str(rownum)
        if sheet["A" + rownum].value == ma:
            tensukien = sheet["B" + rownum].value
            ketqua = sheet["E" + rownum].value
            print(tensukien)
            print(ketqua)
            writeData1(var_app.path_luutamthoi, "Sheet1", 42, 2, tensukien)
            writeData1(var_app.path_luutamthoi, "Sheet1", 43, 2, ketqua)
        rownum = int(rownum)



def write_caseif():
    n = 88
    while (n < 113):
        var_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Utilities"+n+"':\n       caseid_app.caseid_utilities"+n+"(self)\nexcept:\n    module_other_app.close_app()")
        n = int(n)




def write_caseif1():
    n = 1
    while (n < 36):
        var_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   caseid_app.caseid_utilities"+n+"(self)\nexcept:\n     module_other_app.close_app()")
        n = int(n)
        



def write_caseif2():
    n = 84
    while (n < 155):
        var_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("caseid_app.caseid_minitor"+n+"(self='""')")
        n = int(n)
#caseid_app.caseid_minitor83(self="")





def write_caseif3():
    n = 88
    while (n < 113):
        var_app.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Utilities"+n+"':\n       login_app.login.logout_back(self)\n       caseid_app.caseid_utilities"+n+"(self)  \nexcept:\n    module_other_app.close_app()")
        n = int(n)





def write_result_text_try_if_cut(code, eventname, result, path_module, path_text, number_from, number_to, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        var_app.logging.info(check_text[number_from: number_to])
        var_app.logging.info(check_result)
        if check_text[number_from: number_to] == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_cut(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, 10, 15, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_cut_not_fail(code, eventname, result, path_module, path_text, number_from, number_to, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        var_app.logging.info(check_text[number_from: number_to])
        var_app.logging.info(check_result)
        if check_text[number_from: number_to] == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        try:
            no_data = var_app.driver.find_element(By.XPATH, var_app.no_data).text
            writeData(var_app.checklistpath, "Checklist", code, 6, no_data)
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        except:
            pass




def write_result_text_try_if_checked(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).get_attribute("checked")
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_in(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_result in check_text:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_not_fail(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_not_fail_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text != check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
    except:
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_not_fail_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_conten(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text == check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).text
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text != check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_conten_other(code, eventname, result, path_module, path_text, check_result, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_text = var_app.driver.find_element(By.XPATH, path_text).get_attribute("content-desc")
        writeData(var_app.checklistpath, "Checklist", code, 6, check_text)
        var_app.logging.info(check_text)
        if check_text != check_result:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    # write_result_text_try_if_conten_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_app.check_open_car_quickly, "None", "_QuanTri_DsXe_MoXeNhanh.png")








def write_result_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        var_app.driver.find_element(By.XPATH, path_text).is_displayed()
        var_app.logging.info("True")
        writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")



def write_result_not_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_app.driver.implicitly_wait(2)
    var_app.logging.info("--------------")
    var_app.logging.info(path_module)
    var_app.logging.info("Mã - " + code)
    var_app.logging.info("Tên sự kiện - " + eventname)
    var_app.logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var_app.driver.find_element(By.XPATH, path_text).is_displayed()
        var_app.logging.info("False")
        var_app.driver.save_screenshot(var_app.imagepath + code + name_image)
        writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_app.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        var_app.logging.info("True")
        writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
    # chucnangkhac.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_app.check_hide_car, "_QuanTri_DsXe_AnXe.png")

        
        
        
def write_result_excelreport(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_app.logging.info("Check vị trí "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")



def write_result_excelreport_clear_data(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        var_app.logging.info("Check vị trí "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            var_app.logging.info("True")
            writeData(var_app.checklistpath, "Checklist", code, 6, " ")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Pass")
        else:
            var_app.logging.info("False")
            writeData(var_app.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_app.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")

        


