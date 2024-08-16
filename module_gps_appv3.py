import openpyxl
import var_app
import re
import caseid_app
import module_other_app






def ModuleTest():
    var_app.driver.implicitly_wait(7)
    moduletest = ''.join(re.findall(r'\d+', var_app.moduletest))
    print(type(moduletest))
    print(moduletest)
    for i in moduletest:
        print("so", i)
        if i == "0":
            run_all(self='')
        if i == "1":
            login(self='')
        if i == "2":
            minitor(self='')
        if i == "3":
            route(self='')
        if i == "4":
            vehicle(self='')
        if i == "5":
            homepage_toolbar_favorite(self='')
        if i == "6":
            image_camera(self='')
        if i == "7":
            report(self='')
        if i == "8":
            utilitie(self='')




#0
def run_all(self):      #Chạy tất cả
    try:
        caseid_app.caseid_login01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_login20(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_minitor01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor28(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor29(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor30(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor31(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor32(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor33(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor34(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor35(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor36(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor37(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor38(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor39(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor40(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor41(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor42(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor43(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor44(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor45(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor46(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor47(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor48(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor49(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor50(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor51(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor52(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor53(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor54(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor55(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor56(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor57(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor58(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor59(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor60(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor61(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor62(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor63(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor64(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor65(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor66(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor67(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor68(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor69(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor70(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor71(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor72(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor73(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor74(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor75(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor76(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor77(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor78(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor79(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor80(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor81(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor82(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor83(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor84(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor85(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor86(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor87(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor88(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor89(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor90(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor91(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor92(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor93(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor94(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor95(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor96(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor97(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor98(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor99(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor100(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor101(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor102(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor103(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor104(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor105(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor106(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor107(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor108(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor109(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor110(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor111(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor112(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor113(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor114(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor115(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor116(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor117(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor118(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor119(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor120(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor121(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor122(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor123(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor124(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor125(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor126(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor127(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor128(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor129(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor130(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor131(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor132(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor133(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor134(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor135(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor136(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor137(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor138(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor139(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor140(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor141(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor142(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor143(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor144(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor145(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor146(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor147(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor148(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor149(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor150(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor151(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor152(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor153(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor154(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor155(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor156(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor157(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor158(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor159(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor160(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor161(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor162(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor163(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor164(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor165(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor166(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor167(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor168(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor169(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_minitor170(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_route01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_route11(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_vehicle01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle28(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_vehicle29(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_toolbar01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_toolbar28(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_favorite01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_favorite02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_favorite03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_favorite04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_favorite05(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_media01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media28(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media29(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media30(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media31(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media32(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media33(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media34(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media35(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media36(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media37(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media38(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media39(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media40(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media41(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media42(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media43(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media44(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media45(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media46(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media47(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media48(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media49(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media50(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media51(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media52(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media53(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media54(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media55(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media56(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media57(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media58(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media59(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media60(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media61(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media62(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media63(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media64(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media65(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media66(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media67(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_media68(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_report01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report28(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report29(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report30(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report31(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report32(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report33(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report34(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report35(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report36(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report37(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report38(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report39(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report40(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report41(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report42(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report43(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report44(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report45(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report46(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report47(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report48(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report49(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report50(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report51(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report52(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report53(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report54(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report55(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report56(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report57(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report58(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report59(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report60(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report61(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report62(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report63(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report64(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report65(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report66(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report67(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report68(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report69(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report70(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report71(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report72(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report73(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report74(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report75(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report76(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report77(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report78(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report79(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report80(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report81(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report82(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report83(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report84(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report85(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report86(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report87(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report88(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report89(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report90(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report91(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report92(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report93(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report94(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report95(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report96(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report97(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report98(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report99(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report100(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report101(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report102(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report103(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report104(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report105(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report106(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report107(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report108(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report109(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report110(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report111(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report112(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report113(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report114(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report115(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report116(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report117(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report118(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report119(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report120(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report121(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report122(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report123(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report124(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report125(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report126(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report127(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report128(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report129(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report130(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report131(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report132(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report133(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report134(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report135(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report136(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report137(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report138(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report139(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report140(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report141(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report142(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report143(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report144(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report145(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report146(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report147(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report148(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report149(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report150(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report151(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report152(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report153(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report154(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report155(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report156(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report157(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report158(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report159(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report160(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report161(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report162(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report163(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report164(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report165(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report166(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report167(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report168(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_report169(self)
    except:
        module_other_app.close_app()

    try:
        caseid_app.caseid_utilities01(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities02(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities03(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities04(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities05(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities06(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities07(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities08(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities09(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities10(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities11(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities12(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities13(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities14(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities15(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities16(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities17(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities18(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities19(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities20(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities21(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities22(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities23(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities24(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities25(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities26(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities27(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities28(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities29(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities30(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities31(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities32(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities33(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities34(self)
    except:
        module_other_app.close_app()
    try:
        caseid_app.caseid_utilities35(self)
    except:
        module_other_app.close_app()




#1 đăng nhập
def login(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 10
    while (rownum < 34):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Login01':
                        caseid_app.caseid_login01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login02':
                        caseid_app.caseid_login02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login03':
                        caseid_app.caseid_login03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login04':
                        caseid_app.caseid_login04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login05':
                        caseid_app.caseid_login05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login06':
                        caseid_app.caseid_login06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login07':
                        caseid_app.caseid_login07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login08':
                        caseid_app.caseid_login08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login09':
                        caseid_app.caseid_login09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login10':
                        caseid_app.caseid_login10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login11':
                        caseid_app.caseid_login11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login12':
                        caseid_app.caseid_login12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login13':
                        caseid_app.caseid_login13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login14':
                        caseid_app.caseid_login14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login15':
                        caseid_app.caseid_login15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login16':
                        caseid_app.caseid_login16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login17':
                        caseid_app.caseid_login17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login18':
                        caseid_app.caseid_login18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login19':
                        caseid_app.caseid_login19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login20':
                        caseid_app.caseid_login20(self)
                except:
                    module_other_app.close_app()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Login01':
                        caseid_app.caseid_login01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login02':
                        caseid_app.caseid_login02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login03':
                        caseid_app.caseid_login03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login04':
                        caseid_app.caseid_login04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login05':
                        caseid_app.caseid_login05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login06':
                        caseid_app.caseid_login06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login07':
                        caseid_app.caseid_login07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login08':
                        caseid_app.caseid_login08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login09':
                        caseid_app.caseid_login09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login10':
                        caseid_app.caseid_login10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login11':
                        caseid_app.caseid_login11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login12':
                        caseid_app.caseid_login12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login13':
                        caseid_app.caseid_login13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login14':
                        caseid_app.caseid_login14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login15':
                        caseid_app.caseid_login15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login16':
                        caseid_app.caseid_login16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login17':
                        caseid_app.caseid_login17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login18':
                        caseid_app.caseid_login18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login19':
                        caseid_app.caseid_login19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login20':
                        caseid_app.caseid_login20(self)
                except:
                    module_other_app.close_app()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Login01':
                        caseid_app.caseid_login01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login02':
                        caseid_app.caseid_login02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login03':
                        caseid_app.caseid_login03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login04':
                        caseid_app.caseid_login04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login05':
                        caseid_app.caseid_login05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login06':
                        caseid_app.caseid_login06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login07':
                        caseid_app.caseid_login07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login08':
                        caseid_app.caseid_login08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login09':
                        caseid_app.caseid_login09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login10':
                        caseid_app.caseid_login10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login11':
                        caseid_app.caseid_login11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login12':
                        caseid_app.caseid_login12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login13':
                        caseid_app.caseid_login13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login14':
                        caseid_app.caseid_login14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login15':
                        caseid_app.caseid_login15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login16':
                        caseid_app.caseid_login16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login17':
                        caseid_app.caseid_login17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login18':
                        caseid_app.caseid_login18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login19':
                        caseid_app.caseid_login19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login20':
                        caseid_app.caseid_login20(self)
                except:
                    module_other_app.close_app()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Login01':
                        caseid_app.caseid_login01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login02':
                        caseid_app.caseid_login02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login03':
                        caseid_app.caseid_login03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login04':
                        caseid_app.caseid_login04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login05':
                        caseid_app.caseid_login05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login06':
                        caseid_app.caseid_login06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login07':
                        caseid_app.caseid_login07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login08':
                        caseid_app.caseid_login08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login09':
                        caseid_app.caseid_login09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login10':
                        caseid_app.caseid_login10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login11':
                        caseid_app.caseid_login11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login12':
                        caseid_app.caseid_login12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login13':
                        caseid_app.caseid_login13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login14':
                        caseid_app.caseid_login14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login15':
                        caseid_app.caseid_login15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login16':
                        caseid_app.caseid_login16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login17':
                        caseid_app.caseid_login17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login18':
                        caseid_app.caseid_login18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login19':
                        caseid_app.caseid_login19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Login20':
                        caseid_app.caseid_login20(self)
                except:
                    module_other_app.close_app()


#2 Giám sát
def minitor(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 35
    while (rownum < 232):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Minitor01':
                        caseid_app.caseid_minitor01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor02':
                        caseid_app.caseid_minitor02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor03':
                        caseid_app.caseid_minitor03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor04':
                        caseid_app.caseid_minitor04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor05':
                        caseid_app.caseid_minitor05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor06':
                        caseid_app.caseid_minitor06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor07':
                        caseid_app.caseid_minitor07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor08':
                        caseid_app.caseid_minitor08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor09':
                        caseid_app.caseid_minitor09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor10':
                        caseid_app.caseid_minitor10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor11':
                        caseid_app.caseid_minitor11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor12':
                        caseid_app.caseid_minitor12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor13':
                        caseid_app.caseid_minitor13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor14':
                        caseid_app.caseid_minitor14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor15':
                        caseid_app.caseid_minitor15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor16':
                        caseid_app.caseid_minitor16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor17':
                        caseid_app.caseid_minitor17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor18':
                        caseid_app.caseid_minitor18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor19':
                        caseid_app.caseid_minitor19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor20':
                        caseid_app.caseid_minitor20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor21':
                        caseid_app.caseid_minitor21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor22':
                        caseid_app.caseid_minitor22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor23':
                        caseid_app.caseid_minitor23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor24':
                        caseid_app.caseid_minitor24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor25':
                        caseid_app.caseid_minitor25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor26':
                        caseid_app.caseid_minitor26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor27':
                        caseid_app.caseid_minitor27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor28':
                        caseid_app.caseid_minitor28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor29':
                        caseid_app.caseid_minitor29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor30':
                        caseid_app.caseid_minitor30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor31':
                        caseid_app.caseid_minitor31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor32':
                        caseid_app.caseid_minitor32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor33':
                        caseid_app.caseid_minitor33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor34':
                        caseid_app.caseid_minitor34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor35':
                        caseid_app.caseid_minitor35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor36':
                        caseid_app.caseid_minitor36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor37':
                        caseid_app.caseid_minitor37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor38':
                        caseid_app.caseid_minitor38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor39':
                        caseid_app.caseid_minitor39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor40':
                        caseid_app.caseid_minitor40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor41':
                        caseid_app.caseid_minitor41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor42':
                        caseid_app.caseid_minitor42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor43':
                        caseid_app.caseid_minitor43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor44':
                        caseid_app.caseid_minitor44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor45':
                        caseid_app.caseid_minitor45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor46':
                        caseid_app.caseid_minitor46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor47':
                        caseid_app.caseid_minitor47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor48':
                        caseid_app.caseid_minitor48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor49':
                        caseid_app.caseid_minitor49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor50':
                        caseid_app.caseid_minitor50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor51':
                        caseid_app.caseid_minitor51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor52':
                        caseid_app.caseid_minitor52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor53':
                        caseid_app.caseid_minitor53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor54':
                        caseid_app.caseid_minitor54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor55':
                        caseid_app.caseid_minitor55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor56':
                        caseid_app.caseid_minitor56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor57':
                        caseid_app.caseid_minitor57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor58':
                        caseid_app.caseid_minitor58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor59':
                        caseid_app.caseid_minitor59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor60':
                        caseid_app.caseid_minitor60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor61':
                        caseid_app.caseid_minitor61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor62':
                        caseid_app.caseid_minitor62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor63':
                        caseid_app.caseid_minitor63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor64':
                        caseid_app.caseid_minitor64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor65':
                        caseid_app.caseid_minitor65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor66':
                        caseid_app.caseid_minitor66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor67':
                        caseid_app.caseid_minitor67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor68':
                        caseid_app.caseid_minitor68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor69':
                        caseid_app.caseid_minitor69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor70':
                        caseid_app.caseid_minitor70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor71':
                        caseid_app.caseid_minitor71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor72':
                        caseid_app.caseid_minitor72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor73':
                        caseid_app.caseid_minitor73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor74':
                        caseid_app.caseid_minitor74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor75':
                        caseid_app.caseid_minitor75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor76':
                        caseid_app.caseid_minitor76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor77':
                        caseid_app.caseid_minitor77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor78':
                        caseid_app.caseid_minitor78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor79':
                        caseid_app.caseid_minitor79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor80':
                        caseid_app.caseid_minitor80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor81':
                        caseid_app.caseid_minitor81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor82':
                        caseid_app.caseid_minitor82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor83':
                        caseid_app.caseid_minitor83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor84':
                        caseid_app.caseid_minitor84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor85':
                        caseid_app.caseid_minitor85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor86':
                        caseid_app.caseid_minitor86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor87':
                        caseid_app.caseid_minitor87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor88':
                        caseid_app.caseid_minitor88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor89':
                        caseid_app.caseid_minitor89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor90':
                        caseid_app.caseid_minitor90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor91':
                        caseid_app.caseid_minitor91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor92':
                        caseid_app.caseid_minitor92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor93':
                        caseid_app.caseid_minitor93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor94':
                        caseid_app.caseid_minitor94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor95':
                        caseid_app.caseid_minitor95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor96':
                        caseid_app.caseid_minitor96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor97':
                        caseid_app.caseid_minitor97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor98':
                        caseid_app.caseid_minitor98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor99':
                        caseid_app.caseid_minitor99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor100':
                        caseid_app.caseid_minitor100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor101':
                        caseid_app.caseid_minitor101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor102':
                        caseid_app.caseid_minitor102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor103':
                        caseid_app.caseid_minitor103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor104':
                        caseid_app.caseid_minitor104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor105':
                        caseid_app.caseid_minitor105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor106':
                        caseid_app.caseid_minitor106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor107':
                        caseid_app.caseid_minitor107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor108':
                        caseid_app.caseid_minitor108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor109':
                        caseid_app.caseid_minitor109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor110':
                        caseid_app.caseid_minitor110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor111':
                        caseid_app.caseid_minitor111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor112':
                        caseid_app.caseid_minitor112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor113':
                        caseid_app.caseid_minitor113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor114':
                        caseid_app.caseid_minitor114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor115':
                        caseid_app.caseid_minitor115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor116':
                        caseid_app.caseid_minitor116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor117':
                        caseid_app.caseid_minitor117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor118':
                        caseid_app.caseid_minitor118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor119':
                        caseid_app.caseid_minitor119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor120':
                        caseid_app.caseid_minitor120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor121':
                        caseid_app.caseid_minitor121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor122':
                        caseid_app.caseid_minitor122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor123':
                        caseid_app.caseid_minitor123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor124':
                        caseid_app.caseid_minitor124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor125':
                        caseid_app.caseid_minitor125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor126':
                        caseid_app.caseid_minitor126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor127':
                        caseid_app.caseid_minitor127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor128':
                        caseid_app.caseid_minitor128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor129':
                        caseid_app.caseid_minitor129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor130':
                        caseid_app.caseid_minitor130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor131':
                        caseid_app.caseid_minitor131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor132':
                        caseid_app.caseid_minitor132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor133':
                        caseid_app.caseid_minitor133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor134':
                        caseid_app.caseid_minitor134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor135':
                        caseid_app.caseid_minitor135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor136':
                        caseid_app.caseid_minitor136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor137':
                        caseid_app.caseid_minitor137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor138':
                        caseid_app.caseid_minitor138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor139':
                        caseid_app.caseid_minitor139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor140':
                        caseid_app.caseid_minitor140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor141':
                        caseid_app.caseid_minitor141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor142':
                        caseid_app.caseid_minitor142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor143':
                        caseid_app.caseid_minitor143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor144':
                        caseid_app.caseid_minitor144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor145':
                        caseid_app.caseid_minitor145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor146':
                        caseid_app.caseid_minitor146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor147':
                        caseid_app.caseid_minitor147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor148':
                        caseid_app.caseid_minitor148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor149':
                        caseid_app.caseid_minitor149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor150':
                        caseid_app.caseid_minitor150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor151':
                        caseid_app.caseid_minitor151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor152':
                        caseid_app.caseid_minitor152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor153':
                        caseid_app.caseid_minitor153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor154':
                        caseid_app.caseid_minitor154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor155':
                        caseid_app.caseid_minitor155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor156':
                        caseid_app.caseid_minitor156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor157':
                        caseid_app.caseid_minitor157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor158':
                        caseid_app.caseid_minitor158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor159':
                        caseid_app.caseid_minitor159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor160':
                        caseid_app.caseid_minitor160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor161':
                        caseid_app.caseid_minitor161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor162':
                        caseid_app.caseid_minitor162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor163':
                        caseid_app.caseid_minitor163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor164':
                        caseid_app.caseid_minitor164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor165':
                        caseid_app.caseid_minitor165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor166':
                        caseid_app.caseid_minitor166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor167':
                        caseid_app.caseid_minitor167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor168':
                        caseid_app.caseid_minitor168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor169':
                        caseid_app.caseid_minitor169(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor170':
                        caseid_app.caseid_minitor170(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Minitor01':
                        caseid_app.caseid_minitor01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor02':
                        caseid_app.caseid_minitor02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor03':
                        caseid_app.caseid_minitor03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor04':
                        caseid_app.caseid_minitor04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor05':
                        caseid_app.caseid_minitor05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor06':
                        caseid_app.caseid_minitor06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor07':
                        caseid_app.caseid_minitor07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor08':
                        caseid_app.caseid_minitor08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor09':
                        caseid_app.caseid_minitor09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor10':
                        caseid_app.caseid_minitor10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor11':
                        caseid_app.caseid_minitor11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor12':
                        caseid_app.caseid_minitor12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor13':
                        caseid_app.caseid_minitor13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor14':
                        caseid_app.caseid_minitor14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor15':
                        caseid_app.caseid_minitor15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor16':
                        caseid_app.caseid_minitor16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor17':
                        caseid_app.caseid_minitor17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor18':
                        caseid_app.caseid_minitor18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor19':
                        caseid_app.caseid_minitor19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor20':
                        caseid_app.caseid_minitor20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor21':
                        caseid_app.caseid_minitor21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor22':
                        caseid_app.caseid_minitor22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor23':
                        caseid_app.caseid_minitor23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor24':
                        caseid_app.caseid_minitor24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor25':
                        caseid_app.caseid_minitor25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor26':
                        caseid_app.caseid_minitor26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor27':
                        caseid_app.caseid_minitor27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor28':
                        caseid_app.caseid_minitor28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor29':
                        caseid_app.caseid_minitor29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor30':
                        caseid_app.caseid_minitor30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor31':
                        caseid_app.caseid_minitor31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor32':
                        caseid_app.caseid_minitor32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor33':
                        caseid_app.caseid_minitor33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor34':
                        caseid_app.caseid_minitor34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor35':
                        caseid_app.caseid_minitor35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor36':
                        caseid_app.caseid_minitor36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor37':
                        caseid_app.caseid_minitor37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor38':
                        caseid_app.caseid_minitor38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor39':
                        caseid_app.caseid_minitor39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor40':
                        caseid_app.caseid_minitor40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor41':
                        caseid_app.caseid_minitor41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor42':
                        caseid_app.caseid_minitor42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor43':
                        caseid_app.caseid_minitor43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor44':
                        caseid_app.caseid_minitor44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor45':
                        caseid_app.caseid_minitor45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor46':
                        caseid_app.caseid_minitor46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor47':
                        caseid_app.caseid_minitor47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor48':
                        caseid_app.caseid_minitor48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor49':
                        caseid_app.caseid_minitor49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor50':
                        caseid_app.caseid_minitor50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor51':
                        caseid_app.caseid_minitor51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor52':
                        caseid_app.caseid_minitor52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor53':
                        caseid_app.caseid_minitor53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor54':
                        caseid_app.caseid_minitor54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor55':
                        caseid_app.caseid_minitor55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor56':
                        caseid_app.caseid_minitor56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor57':
                        caseid_app.caseid_minitor57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor58':
                        caseid_app.caseid_minitor58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor59':
                        caseid_app.caseid_minitor59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor60':
                        caseid_app.caseid_minitor60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor61':
                        caseid_app.caseid_minitor61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor62':
                        caseid_app.caseid_minitor62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor63':
                        caseid_app.caseid_minitor63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor64':
                        caseid_app.caseid_minitor64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor65':
                        caseid_app.caseid_minitor65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor66':
                        caseid_app.caseid_minitor66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor67':
                        caseid_app.caseid_minitor67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor68':
                        caseid_app.caseid_minitor68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor69':
                        caseid_app.caseid_minitor69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor70':
                        caseid_app.caseid_minitor70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor71':
                        caseid_app.caseid_minitor71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor72':
                        caseid_app.caseid_minitor72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor73':
                        caseid_app.caseid_minitor73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor74':
                        caseid_app.caseid_minitor74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor75':
                        caseid_app.caseid_minitor75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor76':
                        caseid_app.caseid_minitor76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor77':
                        caseid_app.caseid_minitor77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor78':
                        caseid_app.caseid_minitor78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor79':
                        caseid_app.caseid_minitor79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor80':
                        caseid_app.caseid_minitor80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor81':
                        caseid_app.caseid_minitor81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor82':
                        caseid_app.caseid_minitor82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor83':
                        caseid_app.caseid_minitor83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor84':
                        caseid_app.caseid_minitor84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor85':
                        caseid_app.caseid_minitor85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor86':
                        caseid_app.caseid_minitor86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor87':
                        caseid_app.caseid_minitor87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor88':
                        caseid_app.caseid_minitor88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor89':
                        caseid_app.caseid_minitor89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor90':
                        caseid_app.caseid_minitor90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor91':
                        caseid_app.caseid_minitor91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor92':
                        caseid_app.caseid_minitor92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor93':
                        caseid_app.caseid_minitor93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor94':
                        caseid_app.caseid_minitor94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor95':
                        caseid_app.caseid_minitor95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor96':
                        caseid_app.caseid_minitor96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor97':
                        caseid_app.caseid_minitor97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor98':
                        caseid_app.caseid_minitor98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor99':
                        caseid_app.caseid_minitor99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor100':
                        caseid_app.caseid_minitor100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor101':
                        caseid_app.caseid_minitor101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor102':
                        caseid_app.caseid_minitor102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor103':
                        caseid_app.caseid_minitor103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor104':
                        caseid_app.caseid_minitor104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor105':
                        caseid_app.caseid_minitor105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor106':
                        caseid_app.caseid_minitor106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor107':
                        caseid_app.caseid_minitor107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor108':
                        caseid_app.caseid_minitor108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor109':
                        caseid_app.caseid_minitor109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor110':
                        caseid_app.caseid_minitor110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor111':
                        caseid_app.caseid_minitor111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor112':
                        caseid_app.caseid_minitor112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor113':
                        caseid_app.caseid_minitor113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor114':
                        caseid_app.caseid_minitor114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor115':
                        caseid_app.caseid_minitor115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor116':
                        caseid_app.caseid_minitor116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor117':
                        caseid_app.caseid_minitor117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor118':
                        caseid_app.caseid_minitor118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor119':
                        caseid_app.caseid_minitor119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor120':
                        caseid_app.caseid_minitor120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor121':
                        caseid_app.caseid_minitor121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor122':
                        caseid_app.caseid_minitor122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor123':
                        caseid_app.caseid_minitor123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor124':
                        caseid_app.caseid_minitor124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor125':
                        caseid_app.caseid_minitor125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor126':
                        caseid_app.caseid_minitor126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor127':
                        caseid_app.caseid_minitor127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor128':
                        caseid_app.caseid_minitor128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor129':
                        caseid_app.caseid_minitor129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor130':
                        caseid_app.caseid_minitor130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor131':
                        caseid_app.caseid_minitor131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor132':
                        caseid_app.caseid_minitor132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor133':
                        caseid_app.caseid_minitor133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor134':
                        caseid_app.caseid_minitor134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor135':
                        caseid_app.caseid_minitor135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor136':
                        caseid_app.caseid_minitor136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor137':
                        caseid_app.caseid_minitor137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor138':
                        caseid_app.caseid_minitor138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor139':
                        caseid_app.caseid_minitor139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor140':
                        caseid_app.caseid_minitor140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor141':
                        caseid_app.caseid_minitor141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor142':
                        caseid_app.caseid_minitor142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor143':
                        caseid_app.caseid_minitor143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor144':
                        caseid_app.caseid_minitor144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor145':
                        caseid_app.caseid_minitor145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor146':
                        caseid_app.caseid_minitor146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor147':
                        caseid_app.caseid_minitor147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor148':
                        caseid_app.caseid_minitor148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor149':
                        caseid_app.caseid_minitor149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor150':
                        caseid_app.caseid_minitor150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor151':
                        caseid_app.caseid_minitor151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor152':
                        caseid_app.caseid_minitor152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor153':
                        caseid_app.caseid_minitor153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor154':
                        caseid_app.caseid_minitor154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor155':
                        caseid_app.caseid_minitor155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor156':
                        caseid_app.caseid_minitor156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor157':
                        caseid_app.caseid_minitor157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor158':
                        caseid_app.caseid_minitor158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor159':
                        caseid_app.caseid_minitor159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor160':
                        caseid_app.caseid_minitor160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor161':
                        caseid_app.caseid_minitor161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor162':
                        caseid_app.caseid_minitor162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor163':
                        caseid_app.caseid_minitor163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor164':
                        caseid_app.caseid_minitor164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor165':
                        caseid_app.caseid_minitor165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor166':
                        caseid_app.caseid_minitor166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor167':
                        caseid_app.caseid_minitor167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor168':
                        caseid_app.caseid_minitor168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor169':
                        caseid_app.caseid_minitor169(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor170':
                        caseid_app.caseid_minitor170(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Minitor01':
                        caseid_app.caseid_minitor01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor02':
                        caseid_app.caseid_minitor02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor03':
                        caseid_app.caseid_minitor03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor04':
                        caseid_app.caseid_minitor04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor05':
                        caseid_app.caseid_minitor05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor06':
                        caseid_app.caseid_minitor06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor07':
                        caseid_app.caseid_minitor07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor08':
                        caseid_app.caseid_minitor08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor09':
                        caseid_app.caseid_minitor09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor10':
                        caseid_app.caseid_minitor10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor11':
                        caseid_app.caseid_minitor11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor12':
                        caseid_app.caseid_minitor12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor13':
                        caseid_app.caseid_minitor13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor14':
                        caseid_app.caseid_minitor14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor15':
                        caseid_app.caseid_minitor15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor16':
                        caseid_app.caseid_minitor16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor17':
                        caseid_app.caseid_minitor17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor18':
                        caseid_app.caseid_minitor18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor19':
                        caseid_app.caseid_minitor19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor20':
                        caseid_app.caseid_minitor20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor21':
                        caseid_app.caseid_minitor21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor22':
                        caseid_app.caseid_minitor22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor23':
                        caseid_app.caseid_minitor23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor24':
                        caseid_app.caseid_minitor24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor25':
                        caseid_app.caseid_minitor25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor26':
                        caseid_app.caseid_minitor26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor27':
                        caseid_app.caseid_minitor27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor28':
                        caseid_app.caseid_minitor28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor29':
                        caseid_app.caseid_minitor29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor30':
                        caseid_app.caseid_minitor30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor31':
                        caseid_app.caseid_minitor31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor32':
                        caseid_app.caseid_minitor32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor33':
                        caseid_app.caseid_minitor33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor34':
                        caseid_app.caseid_minitor34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor35':
                        caseid_app.caseid_minitor35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor36':
                        caseid_app.caseid_minitor36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor37':
                        caseid_app.caseid_minitor37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor38':
                        caseid_app.caseid_minitor38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor39':
                        caseid_app.caseid_minitor39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor40':
                        caseid_app.caseid_minitor40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor41':
                        caseid_app.caseid_minitor41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor42':
                        caseid_app.caseid_minitor42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor43':
                        caseid_app.caseid_minitor43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor44':
                        caseid_app.caseid_minitor44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor45':
                        caseid_app.caseid_minitor45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor46':
                        caseid_app.caseid_minitor46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor47':
                        caseid_app.caseid_minitor47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor48':
                        caseid_app.caseid_minitor48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor49':
                        caseid_app.caseid_minitor49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor50':
                        caseid_app.caseid_minitor50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor51':
                        caseid_app.caseid_minitor51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor52':
                        caseid_app.caseid_minitor52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor53':
                        caseid_app.caseid_minitor53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor54':
                        caseid_app.caseid_minitor54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor55':
                        caseid_app.caseid_minitor55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor56':
                        caseid_app.caseid_minitor56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor57':
                        caseid_app.caseid_minitor57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor58':
                        caseid_app.caseid_minitor58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor59':
                        caseid_app.caseid_minitor59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor60':
                        caseid_app.caseid_minitor60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor61':
                        caseid_app.caseid_minitor61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor62':
                        caseid_app.caseid_minitor62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor63':
                        caseid_app.caseid_minitor63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor64':
                        caseid_app.caseid_minitor64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor65':
                        caseid_app.caseid_minitor65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor66':
                        caseid_app.caseid_minitor66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor67':
                        caseid_app.caseid_minitor67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor68':
                        caseid_app.caseid_minitor68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor69':
                        caseid_app.caseid_minitor69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor70':
                        caseid_app.caseid_minitor70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor71':
                        caseid_app.caseid_minitor71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor72':
                        caseid_app.caseid_minitor72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor73':
                        caseid_app.caseid_minitor73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor74':
                        caseid_app.caseid_minitor74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor75':
                        caseid_app.caseid_minitor75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor76':
                        caseid_app.caseid_minitor76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor77':
                        caseid_app.caseid_minitor77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor78':
                        caseid_app.caseid_minitor78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor79':
                        caseid_app.caseid_minitor79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor80':
                        caseid_app.caseid_minitor80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor81':
                        caseid_app.caseid_minitor81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor82':
                        caseid_app.caseid_minitor82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor83':
                        caseid_app.caseid_minitor83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor84':
                        caseid_app.caseid_minitor84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor85':
                        caseid_app.caseid_minitor85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor86':
                        caseid_app.caseid_minitor86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor87':
                        caseid_app.caseid_minitor87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor88':
                        caseid_app.caseid_minitor88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor89':
                        caseid_app.caseid_minitor89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor90':
                        caseid_app.caseid_minitor90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor91':
                        caseid_app.caseid_minitor91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor92':
                        caseid_app.caseid_minitor92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor93':
                        caseid_app.caseid_minitor93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor94':
                        caseid_app.caseid_minitor94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor95':
                        caseid_app.caseid_minitor95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor96':
                        caseid_app.caseid_minitor96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor97':
                        caseid_app.caseid_minitor97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor98':
                        caseid_app.caseid_minitor98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor99':
                        caseid_app.caseid_minitor99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor100':
                        caseid_app.caseid_minitor100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor101':
                        caseid_app.caseid_minitor101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor102':
                        caseid_app.caseid_minitor102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor103':
                        caseid_app.caseid_minitor103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor104':
                        caseid_app.caseid_minitor104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor105':
                        caseid_app.caseid_minitor105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor106':
                        caseid_app.caseid_minitor106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor107':
                        caseid_app.caseid_minitor107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor108':
                        caseid_app.caseid_minitor108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor109':
                        caseid_app.caseid_minitor109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor110':
                        caseid_app.caseid_minitor110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor111':
                        caseid_app.caseid_minitor111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor112':
                        caseid_app.caseid_minitor112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor113':
                        caseid_app.caseid_minitor113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor114':
                        caseid_app.caseid_minitor114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor115':
                        caseid_app.caseid_minitor115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor116':
                        caseid_app.caseid_minitor116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor117':
                        caseid_app.caseid_minitor117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor118':
                        caseid_app.caseid_minitor118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor119':
                        caseid_app.caseid_minitor119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor120':
                        caseid_app.caseid_minitor120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor121':
                        caseid_app.caseid_minitor121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor122':
                        caseid_app.caseid_minitor122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor123':
                        caseid_app.caseid_minitor123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor124':
                        caseid_app.caseid_minitor124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor125':
                        caseid_app.caseid_minitor125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor126':
                        caseid_app.caseid_minitor126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor127':
                        caseid_app.caseid_minitor127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor128':
                        caseid_app.caseid_minitor128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor129':
                        caseid_app.caseid_minitor129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor130':
                        caseid_app.caseid_minitor130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor131':
                        caseid_app.caseid_minitor131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor132':
                        caseid_app.caseid_minitor132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor133':
                        caseid_app.caseid_minitor133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor134':
                        caseid_app.caseid_minitor134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor135':
                        caseid_app.caseid_minitor135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor136':
                        caseid_app.caseid_minitor136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor137':
                        caseid_app.caseid_minitor137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor138':
                        caseid_app.caseid_minitor138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor139':
                        caseid_app.caseid_minitor139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor140':
                        caseid_app.caseid_minitor140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor141':
                        caseid_app.caseid_minitor141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor142':
                        caseid_app.caseid_minitor142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor143':
                        caseid_app.caseid_minitor143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor144':
                        caseid_app.caseid_minitor144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor145':
                        caseid_app.caseid_minitor145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor146':
                        caseid_app.caseid_minitor146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor147':
                        caseid_app.caseid_minitor147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor148':
                        caseid_app.caseid_minitor148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor149':
                        caseid_app.caseid_minitor149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor150':
                        caseid_app.caseid_minitor150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor151':
                        caseid_app.caseid_minitor151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor152':
                        caseid_app.caseid_minitor152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor153':
                        caseid_app.caseid_minitor153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor154':
                        caseid_app.caseid_minitor154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor155':
                        caseid_app.caseid_minitor155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor156':
                        caseid_app.caseid_minitor156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor157':
                        caseid_app.caseid_minitor157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor158':
                        caseid_app.caseid_minitor158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor159':
                        caseid_app.caseid_minitor159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor160':
                        caseid_app.caseid_minitor160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor161':
                        caseid_app.caseid_minitor161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor162':
                        caseid_app.caseid_minitor162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor163':
                        caseid_app.caseid_minitor163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor164':
                        caseid_app.caseid_minitor164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor165':
                        caseid_app.caseid_minitor165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor166':
                        caseid_app.caseid_minitor166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor167':
                        caseid_app.caseid_minitor167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor168':
                        caseid_app.caseid_minitor168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor169':
                        caseid_app.caseid_minitor169(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor170':
                        caseid_app.caseid_minitor170(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Minitor01':
                        caseid_app.caseid_minitor01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor02':
                        caseid_app.caseid_minitor02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor03':
                        caseid_app.caseid_minitor03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor04':
                        caseid_app.caseid_minitor04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor05':
                        caseid_app.caseid_minitor05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor06':
                        caseid_app.caseid_minitor06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor07':
                        caseid_app.caseid_minitor07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor08':
                        caseid_app.caseid_minitor08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor09':
                        caseid_app.caseid_minitor09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor10':
                        caseid_app.caseid_minitor10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor11':
                        caseid_app.caseid_minitor11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor12':
                        caseid_app.caseid_minitor12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor13':
                        caseid_app.caseid_minitor13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor14':
                        caseid_app.caseid_minitor14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor15':
                        caseid_app.caseid_minitor15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor16':
                        caseid_app.caseid_minitor16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor17':
                        caseid_app.caseid_minitor17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor18':
                        caseid_app.caseid_minitor18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor19':
                        caseid_app.caseid_minitor19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor20':
                        caseid_app.caseid_minitor20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor21':
                        caseid_app.caseid_minitor21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor22':
                        caseid_app.caseid_minitor22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor23':
                        caseid_app.caseid_minitor23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor24':
                        caseid_app.caseid_minitor24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor25':
                        caseid_app.caseid_minitor25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor26':
                        caseid_app.caseid_minitor26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor27':
                        caseid_app.caseid_minitor27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor28':
                        caseid_app.caseid_minitor28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor29':
                        caseid_app.caseid_minitor29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor30':
                        caseid_app.caseid_minitor30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor31':
                        caseid_app.caseid_minitor31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor32':
                        caseid_app.caseid_minitor32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor33':
                        caseid_app.caseid_minitor33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor34':
                        caseid_app.caseid_minitor34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor35':
                        caseid_app.caseid_minitor35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor36':
                        caseid_app.caseid_minitor36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor37':
                        caseid_app.caseid_minitor37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor38':
                        caseid_app.caseid_minitor38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor39':
                        caseid_app.caseid_minitor39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor40':
                        caseid_app.caseid_minitor40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor41':
                        caseid_app.caseid_minitor41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor42':
                        caseid_app.caseid_minitor42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor43':
                        caseid_app.caseid_minitor43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor44':
                        caseid_app.caseid_minitor44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor45':
                        caseid_app.caseid_minitor45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor46':
                        caseid_app.caseid_minitor46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor47':
                        caseid_app.caseid_minitor47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor48':
                        caseid_app.caseid_minitor48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor49':
                        caseid_app.caseid_minitor49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor50':
                        caseid_app.caseid_minitor50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor51':
                        caseid_app.caseid_minitor51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor52':
                        caseid_app.caseid_minitor52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor53':
                        caseid_app.caseid_minitor53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor54':
                        caseid_app.caseid_minitor54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor55':
                        caseid_app.caseid_minitor55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor56':
                        caseid_app.caseid_minitor56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor57':
                        caseid_app.caseid_minitor57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor58':
                        caseid_app.caseid_minitor58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor59':
                        caseid_app.caseid_minitor59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor60':
                        caseid_app.caseid_minitor60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor61':
                        caseid_app.caseid_minitor61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor62':
                        caseid_app.caseid_minitor62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor63':
                        caseid_app.caseid_minitor63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor64':
                        caseid_app.caseid_minitor64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor65':
                        caseid_app.caseid_minitor65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor66':
                        caseid_app.caseid_minitor66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor67':
                        caseid_app.caseid_minitor67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor68':
                        caseid_app.caseid_minitor68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor69':
                        caseid_app.caseid_minitor69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor70':
                        caseid_app.caseid_minitor70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor71':
                        caseid_app.caseid_minitor71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor72':
                        caseid_app.caseid_minitor72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor73':
                        caseid_app.caseid_minitor73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor74':
                        caseid_app.caseid_minitor74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor75':
                        caseid_app.caseid_minitor75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor76':
                        caseid_app.caseid_minitor76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor77':
                        caseid_app.caseid_minitor77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor78':
                        caseid_app.caseid_minitor78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor79':
                        caseid_app.caseid_minitor79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor80':
                        caseid_app.caseid_minitor80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor81':
                        caseid_app.caseid_minitor81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor82':
                        caseid_app.caseid_minitor82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor83':
                        caseid_app.caseid_minitor83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor84':
                        caseid_app.caseid_minitor84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor85':
                        caseid_app.caseid_minitor85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor86':
                        caseid_app.caseid_minitor86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor87':
                        caseid_app.caseid_minitor87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor88':
                        caseid_app.caseid_minitor88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor89':
                        caseid_app.caseid_minitor89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor90':
                        caseid_app.caseid_minitor90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor91':
                        caseid_app.caseid_minitor91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor92':
                        caseid_app.caseid_minitor92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor93':
                        caseid_app.caseid_minitor93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor94':
                        caseid_app.caseid_minitor94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor95':
                        caseid_app.caseid_minitor95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor96':
                        caseid_app.caseid_minitor96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor97':
                        caseid_app.caseid_minitor97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor98':
                        caseid_app.caseid_minitor98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor99':
                        caseid_app.caseid_minitor99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor100':
                        caseid_app.caseid_minitor100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor101':
                        caseid_app.caseid_minitor101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor102':
                        caseid_app.caseid_minitor102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor103':
                        caseid_app.caseid_minitor103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor104':
                        caseid_app.caseid_minitor104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor105':
                        caseid_app.caseid_minitor105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor106':
                        caseid_app.caseid_minitor106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor107':
                        caseid_app.caseid_minitor107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor108':
                        caseid_app.caseid_minitor108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor109':
                        caseid_app.caseid_minitor109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor110':
                        caseid_app.caseid_minitor110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor111':
                        caseid_app.caseid_minitor111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor112':
                        caseid_app.caseid_minitor112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor113':
                        caseid_app.caseid_minitor113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor114':
                        caseid_app.caseid_minitor114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor115':
                        caseid_app.caseid_minitor115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor116':
                        caseid_app.caseid_minitor116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor117':
                        caseid_app.caseid_minitor117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor118':
                        caseid_app.caseid_minitor118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor119':
                        caseid_app.caseid_minitor119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor120':
                        caseid_app.caseid_minitor120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor121':
                        caseid_app.caseid_minitor121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor122':
                        caseid_app.caseid_minitor122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor123':
                        caseid_app.caseid_minitor123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor124':
                        caseid_app.caseid_minitor124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor125':
                        caseid_app.caseid_minitor125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor126':
                        caseid_app.caseid_minitor126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor127':
                        caseid_app.caseid_minitor127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor128':
                        caseid_app.caseid_minitor128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor129':
                        caseid_app.caseid_minitor129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor130':
                        caseid_app.caseid_minitor130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor131':
                        caseid_app.caseid_minitor131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor132':
                        caseid_app.caseid_minitor132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor133':
                        caseid_app.caseid_minitor133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor134':
                        caseid_app.caseid_minitor134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor135':
                        caseid_app.caseid_minitor135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor136':
                        caseid_app.caseid_minitor136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor137':
                        caseid_app.caseid_minitor137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor138':
                        caseid_app.caseid_minitor138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor139':
                        caseid_app.caseid_minitor139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor140':
                        caseid_app.caseid_minitor140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor141':
                        caseid_app.caseid_minitor141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor142':
                        caseid_app.caseid_minitor142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor143':
                        caseid_app.caseid_minitor143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor144':
                        caseid_app.caseid_minitor144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor145':
                        caseid_app.caseid_minitor145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor146':
                        caseid_app.caseid_minitor146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor147':
                        caseid_app.caseid_minitor147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor148':
                        caseid_app.caseid_minitor148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor149':
                        caseid_app.caseid_minitor149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor150':
                        caseid_app.caseid_minitor150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor151':
                        caseid_app.caseid_minitor151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor152':
                        caseid_app.caseid_minitor152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor153':
                        caseid_app.caseid_minitor153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor154':
                        caseid_app.caseid_minitor154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor155':
                        caseid_app.caseid_minitor155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor156':
                        caseid_app.caseid_minitor156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor157':
                        caseid_app.caseid_minitor157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor158':
                        caseid_app.caseid_minitor158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor159':
                        caseid_app.caseid_minitor159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor160':
                        caseid_app.caseid_minitor160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor161':
                        caseid_app.caseid_minitor161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor162':
                        caseid_app.caseid_minitor162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor163':
                        caseid_app.caseid_minitor163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor164':
                        caseid_app.caseid_minitor164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor165':
                        caseid_app.caseid_minitor165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor166':
                        caseid_app.caseid_minitor166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor167':
                        caseid_app.caseid_minitor167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor168':
                        caseid_app.caseid_minitor168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor169':
                        caseid_app.caseid_minitor169(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Minitor170':
                        caseid_app.caseid_minitor170(self)
                except:
                    module_other_app.close_app()


#3.Lộ trình
def route(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 233
    while (rownum < 245):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Route01':
                        caseid_app.caseid_route01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route02':
                        caseid_app.caseid_route02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route03':
                        caseid_app.caseid_route03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route04':
                        caseid_app.caseid_route04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route05':
                        caseid_app.caseid_route05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route06':
                        caseid_app.caseid_route06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route07':
                        caseid_app.caseid_route07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route08':
                        caseid_app.caseid_route08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route09':
                        caseid_app.caseid_route09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route10':
                        caseid_app.caseid_route10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route11':
                        caseid_app.caseid_route11(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Route01':
                        caseid_app.caseid_route01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route02':
                        caseid_app.caseid_route02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route03':
                        caseid_app.caseid_route03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route04':
                        caseid_app.caseid_route04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route05':
                        caseid_app.caseid_route05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route06':
                        caseid_app.caseid_route06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route07':
                        caseid_app.caseid_route07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route08':
                        caseid_app.caseid_route08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route09':
                        caseid_app.caseid_route09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route10':
                        caseid_app.caseid_route10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route11':
                        caseid_app.caseid_route11(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Route01':
                        caseid_app.caseid_route01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route02':
                        caseid_app.caseid_route02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route03':
                        caseid_app.caseid_route03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route04':
                        caseid_app.caseid_route04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route05':
                        caseid_app.caseid_route05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route06':
                        caseid_app.caseid_route06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route07':
                        caseid_app.caseid_route07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route08':
                        caseid_app.caseid_route08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route09':
                        caseid_app.caseid_route09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route10':
                        caseid_app.caseid_route10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route11':
                        caseid_app.caseid_route11(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Route01':
                        caseid_app.caseid_route01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route02':
                        caseid_app.caseid_route02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route03':
                        caseid_app.caseid_route03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route04':
                        caseid_app.caseid_route04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route05':
                        caseid_app.caseid_route05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route06':
                        caseid_app.caseid_route06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route07':
                        caseid_app.caseid_route07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route08':
                        caseid_app.caseid_route08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route09':
                        caseid_app.caseid_route09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route10':
                        caseid_app.caseid_route10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Route11':
                        caseid_app.caseid_route11(self)
                except:
                    module_other_app.close_app()


#4. Phương tiện
def vehicle(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 246
    while (rownum < 284):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Vehicle01':
                        caseid_app.caseid_vehicle01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle02':
                        caseid_app.caseid_vehicle02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle03':
                        caseid_app.caseid_vehicle03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle04':
                        caseid_app.caseid_vehicle04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle05':
                        caseid_app.caseid_vehicle05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle06':
                        caseid_app.caseid_vehicle06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle07':
                        caseid_app.caseid_vehicle07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle08':
                        caseid_app.caseid_vehicle08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle09':
                        caseid_app.caseid_vehicle09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle10':
                        caseid_app.caseid_vehicle10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle11':
                        caseid_app.caseid_vehicle11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle12':
                        caseid_app.caseid_vehicle12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle13':
                        caseid_app.caseid_vehicle13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle14':
                        caseid_app.caseid_vehicle14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle15':
                        caseid_app.caseid_vehicle15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle16':
                        caseid_app.caseid_vehicle16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle17':
                        caseid_app.caseid_vehicle17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle18':
                        caseid_app.caseid_vehicle18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle19':
                        caseid_app.caseid_vehicle19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle20':
                        caseid_app.caseid_vehicle20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle21':
                        caseid_app.caseid_vehicle21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle22':
                        caseid_app.caseid_vehicle22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle23':
                        caseid_app.caseid_vehicle23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle24':
                        caseid_app.caseid_vehicle24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle25':
                        caseid_app.caseid_vehicle25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle26':
                        caseid_app.caseid_vehicle26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle27':
                        caseid_app.caseid_vehicle27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle28':
                        caseid_app.caseid_vehicle28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle29':
                        caseid_app.caseid_vehicle29(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Vehicle01':
                        caseid_app.caseid_vehicle01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle02':
                        caseid_app.caseid_vehicle02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle03':
                        caseid_app.caseid_vehicle03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle04':
                        caseid_app.caseid_vehicle04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle05':
                        caseid_app.caseid_vehicle05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle06':
                        caseid_app.caseid_vehicle06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle07':
                        caseid_app.caseid_vehicle07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle08':
                        caseid_app.caseid_vehicle08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle09':
                        caseid_app.caseid_vehicle09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle10':
                        caseid_app.caseid_vehicle10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle11':
                        caseid_app.caseid_vehicle11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle12':
                        caseid_app.caseid_vehicle12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle13':
                        caseid_app.caseid_vehicle13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle14':
                        caseid_app.caseid_vehicle14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle15':
                        caseid_app.caseid_vehicle15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle16':
                        caseid_app.caseid_vehicle16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle17':
                        caseid_app.caseid_vehicle17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle18':
                        caseid_app.caseid_vehicle18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle19':
                        caseid_app.caseid_vehicle19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle20':
                        caseid_app.caseid_vehicle20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle21':
                        caseid_app.caseid_vehicle21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle22':
                        caseid_app.caseid_vehicle22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle23':
                        caseid_app.caseid_vehicle23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle24':
                        caseid_app.caseid_vehicle24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle25':
                        caseid_app.caseid_vehicle25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle26':
                        caseid_app.caseid_vehicle26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle27':
                        caseid_app.caseid_vehicle27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle28':
                        caseid_app.caseid_vehicle28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle29':
                        caseid_app.caseid_vehicle29(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Vehicle01':
                        caseid_app.caseid_vehicle01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle02':
                        caseid_app.caseid_vehicle02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle03':
                        caseid_app.caseid_vehicle03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle04':
                        caseid_app.caseid_vehicle04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle05':
                        caseid_app.caseid_vehicle05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle06':
                        caseid_app.caseid_vehicle06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle07':
                        caseid_app.caseid_vehicle07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle08':
                        caseid_app.caseid_vehicle08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle09':
                        caseid_app.caseid_vehicle09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle10':
                        caseid_app.caseid_vehicle10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle11':
                        caseid_app.caseid_vehicle11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle12':
                        caseid_app.caseid_vehicle12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle13':
                        caseid_app.caseid_vehicle13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle14':
                        caseid_app.caseid_vehicle14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle15':
                        caseid_app.caseid_vehicle15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle16':
                        caseid_app.caseid_vehicle16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle17':
                        caseid_app.caseid_vehicle17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle18':
                        caseid_app.caseid_vehicle18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle19':
                        caseid_app.caseid_vehicle19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle20':
                        caseid_app.caseid_vehicle20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle21':
                        caseid_app.caseid_vehicle21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle22':
                        caseid_app.caseid_vehicle22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle23':
                        caseid_app.caseid_vehicle23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle24':
                        caseid_app.caseid_vehicle24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle25':
                        caseid_app.caseid_vehicle25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle26':
                        caseid_app.caseid_vehicle26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle27':
                        caseid_app.caseid_vehicle27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle28':
                        caseid_app.caseid_vehicle28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle29':
                        caseid_app.caseid_vehicle29(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Vehicle01':
                        caseid_app.caseid_vehicle01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle02':
                        caseid_app.caseid_vehicle02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle03':
                        caseid_app.caseid_vehicle03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle04':
                        caseid_app.caseid_vehicle04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle05':
                        caseid_app.caseid_vehicle05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle06':
                        caseid_app.caseid_vehicle06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle07':
                        caseid_app.caseid_vehicle07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle08':
                        caseid_app.caseid_vehicle08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle09':
                        caseid_app.caseid_vehicle09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle10':
                        caseid_app.caseid_vehicle10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle11':
                        caseid_app.caseid_vehicle11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle12':
                        caseid_app.caseid_vehicle12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle13':
                        caseid_app.caseid_vehicle13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle14':
                        caseid_app.caseid_vehicle14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle15':
                        caseid_app.caseid_vehicle15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle16':
                        caseid_app.caseid_vehicle16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle17':
                        caseid_app.caseid_vehicle17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle18':
                        caseid_app.caseid_vehicle18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle19':
                        caseid_app.caseid_vehicle19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle20':
                        caseid_app.caseid_vehicle20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle21':
                        caseid_app.caseid_vehicle21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle22':
                        caseid_app.caseid_vehicle22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle23':
                        caseid_app.caseid_vehicle23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle24':
                        caseid_app.caseid_vehicle24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle25':
                        caseid_app.caseid_vehicle25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle26':
                        caseid_app.caseid_vehicle26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle27':
                        caseid_app.caseid_vehicle27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle28':
                        caseid_app.caseid_vehicle28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Vehicle29':
                        caseid_app.caseid_vehicle29(self)
                except:
                    module_other_app.close_app()



#5 Trang chủ - thanh công cụ/yêu thích
def homepage_toolbar_favorite(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 286
    while (rownum < 338):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Toolbar01':
                        caseid_app.caseid_toolbar01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar02':
                        caseid_app.caseid_toolbar02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar03':
                        caseid_app.caseid_toolbar03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar04':
                        caseid_app.caseid_toolbar04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar05':
                        caseid_app.caseid_toolbar05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar6':
                        caseid_app.caseid_toolbar06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar7':
                        caseid_app.caseid_toolbar07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar08':
                        caseid_app.caseid_toolbar08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar09':
                        caseid_app.caseid_toolbar09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar10':
                        caseid_app.caseid_toolbar10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar11':
                        caseid_app.caseid_toolbar11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar12':
                        caseid_app.caseid_toolbar12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar13':
                        caseid_app.caseid_toolbar13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar14':
                        caseid_app.caseid_toolbar14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar15':
                        caseid_app.caseid_toolbar15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar16':
                        caseid_app.caseid_toolbar16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar17':
                        caseid_app.caseid_toolbar17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar18':
                        caseid_app.caseid_toolbar18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar19':
                        caseid_app.caseid_toolbar19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar20':
                        caseid_app.caseid_toolbar20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar21':
                        caseid_app.caseid_toolbar21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar22':
                        caseid_app.caseid_toolbar22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar23':
                        caseid_app.caseid_toolbar23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar24':
                        caseid_app.caseid_toolbar24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar25':
                        caseid_app.caseid_toolbar25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar26':
                        caseid_app.caseid_toolbar26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar27':
                        caseid_app.caseid_toolbar27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar28':
                        caseid_app.caseid_toolbar28(self)
                except:
                    module_other_app.close_app()

                try:
                    if case == 'Favorite01':
                        caseid_app.caseid_favorite01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite02':
                        caseid_app.caseid_favorite02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite03':
                        caseid_app.caseid_favorite03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite04':
                        caseid_app.caseid_favorite04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite05':
                        caseid_app.caseid_favorite05(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Toolbar01':
                        caseid_app.caseid_toolbar01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar02':
                        caseid_app.caseid_toolbar02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar03':
                        caseid_app.caseid_toolbar03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar04':
                        caseid_app.caseid_toolbar04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar05':
                        caseid_app.caseid_toolbar05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar6':
                        caseid_app.caseid_toolbar06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar7':
                        caseid_app.caseid_toolbar07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar08':
                        caseid_app.caseid_toolbar08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar09':
                        caseid_app.caseid_toolbar09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar10':
                        caseid_app.caseid_toolbar10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar11':
                        caseid_app.caseid_toolbar11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar12':
                        caseid_app.caseid_toolbar12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar13':
                        caseid_app.caseid_toolbar13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar14':
                        caseid_app.caseid_toolbar14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar15':
                        caseid_app.caseid_toolbar15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar16':
                        caseid_app.caseid_toolbar16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar17':
                        caseid_app.caseid_toolbar17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar18':
                        caseid_app.caseid_toolbar18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar19':
                        caseid_app.caseid_toolbar19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar20':
                        caseid_app.caseid_toolbar20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar21':
                        caseid_app.caseid_toolbar21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar22':
                        caseid_app.caseid_toolbar22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar23':
                        caseid_app.caseid_toolbar23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar24':
                        caseid_app.caseid_toolbar24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar25':
                        caseid_app.caseid_toolbar25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar26':
                        caseid_app.caseid_toolbar26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar27':
                        caseid_app.caseid_toolbar27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar28':
                        caseid_app.caseid_toolbar28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite01':
                        caseid_app.caseid_favorite01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite02':
                        caseid_app.caseid_favorite02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite03':
                        caseid_app.caseid_favorite03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite04':
                        caseid_app.caseid_favorite04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite05':
                        caseid_app.caseid_favorite05(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Toolbar01':
                        caseid_app.caseid_toolbar01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar02':
                        caseid_app.caseid_toolbar02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar03':
                        caseid_app.caseid_toolbar03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar04':
                        caseid_app.caseid_toolbar04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar05':
                        caseid_app.caseid_toolbar05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar6':
                        caseid_app.caseid_toolbar06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar7':
                        caseid_app.caseid_toolbar07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar08':
                        caseid_app.caseid_toolbar08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar09':
                        caseid_app.caseid_toolbar09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar10':
                        caseid_app.caseid_toolbar10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar11':
                        caseid_app.caseid_toolbar11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar12':
                        caseid_app.caseid_toolbar12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar13':
                        caseid_app.caseid_toolbar13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar14':
                        caseid_app.caseid_toolbar14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar15':
                        caseid_app.caseid_toolbar15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar16':
                        caseid_app.caseid_toolbar16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar17':
                        caseid_app.caseid_toolbar17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar18':
                        caseid_app.caseid_toolbar18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar19':
                        caseid_app.caseid_toolbar19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar20':
                        caseid_app.caseid_toolbar20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar21':
                        caseid_app.caseid_toolbar21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar22':
                        caseid_app.caseid_toolbar22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar23':
                        caseid_app.caseid_toolbar23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar24':
                        caseid_app.caseid_toolbar24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar25':
                        caseid_app.caseid_toolbar25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar26':
                        caseid_app.caseid_toolbar26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar27':
                        caseid_app.caseid_toolbar27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar28':
                        caseid_app.caseid_toolbar28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite01':
                        caseid_app.caseid_favorite01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite02':
                        caseid_app.caseid_favorite02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite03':
                        caseid_app.caseid_favorite03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite04':
                        caseid_app.caseid_favorite04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite05':
                        caseid_app.caseid_favorite05(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Toolbar01':
                        caseid_app.caseid_toolbar01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar02':
                        caseid_app.caseid_toolbar02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar03':
                        caseid_app.caseid_toolbar03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar04':
                        caseid_app.caseid_toolbar04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar05':
                        caseid_app.caseid_toolbar05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar6':
                        caseid_app.caseid_toolbar06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar7':
                        caseid_app.caseid_toolbar07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar08':
                        caseid_app.caseid_toolbar08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar09':
                        caseid_app.caseid_toolbar09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar10':
                        caseid_app.caseid_toolbar10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar11':
                        caseid_app.caseid_toolbar11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar12':
                        caseid_app.caseid_toolbar12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar13':
                        caseid_app.caseid_toolbar13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar14':
                        caseid_app.caseid_toolbar14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar15':
                        caseid_app.caseid_toolbar15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar16':
                        caseid_app.caseid_toolbar16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar17':
                        caseid_app.caseid_toolbar17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar18':
                        caseid_app.caseid_toolbar18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar19':
                        caseid_app.caseid_toolbar19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar20':
                        caseid_app.caseid_toolbar20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar21':
                        caseid_app.caseid_toolbar21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar22':
                        caseid_app.caseid_toolbar22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar23':
                        caseid_app.caseid_toolbar23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar24':
                        caseid_app.caseid_toolbar24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar25':
                        caseid_app.caseid_toolbar25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar26':
                        caseid_app.caseid_toolbar26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar27':
                        caseid_app.caseid_toolbar27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Toolbar28':
                        caseid_app.caseid_toolbar28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite01':
                        caseid_app.caseid_favorite01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite02':
                        caseid_app.caseid_favorite02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite03':
                        caseid_app.caseid_favorite03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite04':
                        caseid_app.caseid_favorite04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Favorite05':
                        caseid_app.caseid_favorite05(self)
                except:
                    module_other_app.close_app()



#6 Hình ảnh -Camera
def image_camera(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 339
    while (rownum < 433):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Media01':
                        caseid_app.caseid_media01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media02':
                        caseid_app.caseid_media02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media03':
                        caseid_app.caseid_media03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media04':
                        caseid_app.caseid_media04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media05':
                        caseid_app.caseid_media05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media06':
                        caseid_app.caseid_media06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media07':
                        caseid_app.caseid_media07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media08':
                        caseid_app.caseid_media08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media09':
                        caseid_app.caseid_media09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media10':
                        caseid_app.caseid_media10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media11':
                        caseid_app.caseid_media11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media12':
                        caseid_app.caseid_media12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media13':
                        caseid_app.caseid_media13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media14':
                        caseid_app.caseid_media14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media15':
                        caseid_app.caseid_media15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media16':
                        caseid_app.caseid_media16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media17':
                        caseid_app.caseid_media17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media18':
                        caseid_app.caseid_media18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media19':
                        caseid_app.caseid_media19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media20':
                        caseid_app.caseid_media20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media21':
                        caseid_app.caseid_media21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media22':
                        caseid_app.caseid_media22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media23':
                        caseid_app.caseid_media23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media24':
                        caseid_app.caseid_media24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media25':
                        caseid_app.caseid_media25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media26':
                        caseid_app.caseid_media26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media27':
                        caseid_app.caseid_media27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media28':
                        caseid_app.caseid_media28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media29':
                        caseid_app.caseid_media29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media30':
                        caseid_app.caseid_media30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media31':
                        caseid_app.caseid_media31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media32':
                        caseid_app.caseid_media32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media33':
                        caseid_app.caseid_media33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media34':
                        caseid_app.caseid_media34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media35':
                        caseid_app.caseid_media35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media36':
                        caseid_app.caseid_media36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media37':
                        caseid_app.caseid_media37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media38':
                        caseid_app.caseid_media38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media39':
                        caseid_app.caseid_media39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media40':
                        caseid_app.caseid_media40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media41':
                        caseid_app.caseid_media41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media42':
                        caseid_app.caseid_media42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media43':
                        caseid_app.caseid_media43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media44':
                        caseid_app.caseid_media44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media45':
                        caseid_app.caseid_media45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media46':
                        caseid_app.caseid_media46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media47':
                        caseid_app.caseid_media47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media48':
                        caseid_app.caseid_media48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media49':
                        caseid_app.caseid_media49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media50':
                        caseid_app.caseid_media50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media51':
                        caseid_app.caseid_media51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media52':
                        caseid_app.caseid_media52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media53':
                        caseid_app.caseid_media53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media54':
                        caseid_app.caseid_media54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media55':
                        caseid_app.caseid_media55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media56':
                        caseid_app.caseid_media56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media57':
                        caseid_app.caseid_media57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media58':
                        caseid_app.caseid_media58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media59':
                        caseid_app.caseid_media59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media60':
                        caseid_app.caseid_media60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media61':
                        caseid_app.caseid_media61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media62':
                        caseid_app.caseid_media62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media63':
                        caseid_app.caseid_media63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media64':
                        caseid_app.caseid_media64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media65':
                        caseid_app.caseid_media65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media66':
                        caseid_app.caseid_media66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media67':
                        caseid_app.caseid_media67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media68':
                        caseid_app.caseid_media68(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Media01':
                        caseid_app.caseid_media01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media02':
                        caseid_app.caseid_media02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media03':
                        caseid_app.caseid_media03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media04':
                        caseid_app.caseid_media04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media05':
                        caseid_app.caseid_media05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media06':
                        caseid_app.caseid_media06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media07':
                        caseid_app.caseid_media07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media08':
                        caseid_app.caseid_media08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media09':
                        caseid_app.caseid_media09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media10':
                        caseid_app.caseid_media10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media11':
                        caseid_app.caseid_media11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media12':
                        caseid_app.caseid_media12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media13':
                        caseid_app.caseid_media13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media14':
                        caseid_app.caseid_media14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media15':
                        caseid_app.caseid_media15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media16':
                        caseid_app.caseid_media16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media17':
                        caseid_app.caseid_media17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media18':
                        caseid_app.caseid_media18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media19':
                        caseid_app.caseid_media19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media20':
                        caseid_app.caseid_media20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media21':
                        caseid_app.caseid_media21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media22':
                        caseid_app.caseid_media22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media23':
                        caseid_app.caseid_media23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media24':
                        caseid_app.caseid_media24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media25':
                        caseid_app.caseid_media25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media26':
                        caseid_app.caseid_media26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media27':
                        caseid_app.caseid_media27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media28':
                        caseid_app.caseid_media28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media29':
                        caseid_app.caseid_media29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media30':
                        caseid_app.caseid_media30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media31':
                        caseid_app.caseid_media31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media32':
                        caseid_app.caseid_media32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media33':
                        caseid_app.caseid_media33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media34':
                        caseid_app.caseid_media34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media35':
                        caseid_app.caseid_media35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media36':
                        caseid_app.caseid_media36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media37':
                        caseid_app.caseid_media37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media38':
                        caseid_app.caseid_media38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media39':
                        caseid_app.caseid_media39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media40':
                        caseid_app.caseid_media40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media41':
                        caseid_app.caseid_media41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media42':
                        caseid_app.caseid_media42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media43':
                        caseid_app.caseid_media43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media44':
                        caseid_app.caseid_media44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media45':
                        caseid_app.caseid_media45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media46':
                        caseid_app.caseid_media46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media47':
                        caseid_app.caseid_media47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media48':
                        caseid_app.caseid_media48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media49':
                        caseid_app.caseid_media49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media50':
                        caseid_app.caseid_media50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media51':
                        caseid_app.caseid_media51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media52':
                        caseid_app.caseid_media52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media53':
                        caseid_app.caseid_media53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media54':
                        caseid_app.caseid_media54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media55':
                        caseid_app.caseid_media55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media56':
                        caseid_app.caseid_media56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media57':
                        caseid_app.caseid_media57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media58':
                        caseid_app.caseid_media58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media59':
                        caseid_app.caseid_media59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media60':
                        caseid_app.caseid_media60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media61':
                        caseid_app.caseid_media61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media62':
                        caseid_app.caseid_media62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media63':
                        caseid_app.caseid_media63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media64':
                        caseid_app.caseid_media64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media65':
                        caseid_app.caseid_media65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media66':
                        caseid_app.caseid_media66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media67':
                        caseid_app.caseid_media67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media68':
                        caseid_app.caseid_media68(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Media01':
                        caseid_app.caseid_media01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media02':
                        caseid_app.caseid_media02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media03':
                        caseid_app.caseid_media03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media04':
                        caseid_app.caseid_media04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media05':
                        caseid_app.caseid_media05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media06':
                        caseid_app.caseid_media06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media07':
                        caseid_app.caseid_media07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media08':
                        caseid_app.caseid_media08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media09':
                        caseid_app.caseid_media09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media10':
                        caseid_app.caseid_media10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media11':
                        caseid_app.caseid_media11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media12':
                        caseid_app.caseid_media12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media13':
                        caseid_app.caseid_media13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media14':
                        caseid_app.caseid_media14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media15':
                        caseid_app.caseid_media15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media16':
                        caseid_app.caseid_media16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media17':
                        caseid_app.caseid_media17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media18':
                        caseid_app.caseid_media18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media19':
                        caseid_app.caseid_media19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media20':
                        caseid_app.caseid_media20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media21':
                        caseid_app.caseid_media21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media22':
                        caseid_app.caseid_media22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media23':
                        caseid_app.caseid_media23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media24':
                        caseid_app.caseid_media24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media25':
                        caseid_app.caseid_media25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media26':
                        caseid_app.caseid_media26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media27':
                        caseid_app.caseid_media27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media28':
                        caseid_app.caseid_media28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media29':
                        caseid_app.caseid_media29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media30':
                        caseid_app.caseid_media30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media31':
                        caseid_app.caseid_media31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media32':
                        caseid_app.caseid_media32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media33':
                        caseid_app.caseid_media33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media34':
                        caseid_app.caseid_media34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media35':
                        caseid_app.caseid_media35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media36':
                        caseid_app.caseid_media36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media37':
                        caseid_app.caseid_media37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media38':
                        caseid_app.caseid_media38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media39':
                        caseid_app.caseid_media39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media40':
                        caseid_app.caseid_media40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media41':
                        caseid_app.caseid_media41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media42':
                        caseid_app.caseid_media42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media43':
                        caseid_app.caseid_media43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media44':
                        caseid_app.caseid_media44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media45':
                        caseid_app.caseid_media45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media46':
                        caseid_app.caseid_media46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media47':
                        caseid_app.caseid_media47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media48':
                        caseid_app.caseid_media48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media49':
                        caseid_app.caseid_media49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media50':
                        caseid_app.caseid_media50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media51':
                        caseid_app.caseid_media51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media52':
                        caseid_app.caseid_media52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media53':
                        caseid_app.caseid_media53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media54':
                        caseid_app.caseid_media54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media55':
                        caseid_app.caseid_media55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media56':
                        caseid_app.caseid_media56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media57':
                        caseid_app.caseid_media57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media58':
                        caseid_app.caseid_media58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media59':
                        caseid_app.caseid_media59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media60':
                        caseid_app.caseid_media60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media61':
                        caseid_app.caseid_media61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media62':
                        caseid_app.caseid_media62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media63':
                        caseid_app.caseid_media63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media64':
                        caseid_app.caseid_media64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media65':
                        caseid_app.caseid_media65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media66':
                        caseid_app.caseid_media66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media67':
                        caseid_app.caseid_media67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media68':
                        caseid_app.caseid_media68(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Media01':
                        caseid_app.caseid_media01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media02':
                        caseid_app.caseid_media02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media03':
                        caseid_app.caseid_media03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media04':
                        caseid_app.caseid_media04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media05':
                        caseid_app.caseid_media05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media06':
                        caseid_app.caseid_media06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media07':
                        caseid_app.caseid_media07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media08':
                        caseid_app.caseid_media08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media09':
                        caseid_app.caseid_media09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media10':
                        caseid_app.caseid_media10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media11':
                        caseid_app.caseid_media11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media12':
                        caseid_app.caseid_media12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media13':
                        caseid_app.caseid_media13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media14':
                        caseid_app.caseid_media14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media15':
                        caseid_app.caseid_media15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media16':
                        caseid_app.caseid_media16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media17':
                        caseid_app.caseid_media17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media18':
                        caseid_app.caseid_media18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media19':
                        caseid_app.caseid_media19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media20':
                        caseid_app.caseid_media20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media21':
                        caseid_app.caseid_media21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media22':
                        caseid_app.caseid_media22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media23':
                        caseid_app.caseid_media23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media24':
                        caseid_app.caseid_media24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media25':
                        caseid_app.caseid_media25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media26':
                        caseid_app.caseid_media26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media27':
                        caseid_app.caseid_media27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media28':
                        caseid_app.caseid_media28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media29':
                        caseid_app.caseid_media29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media30':
                        caseid_app.caseid_media30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media31':
                        caseid_app.caseid_media31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media32':
                        caseid_app.caseid_media32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media33':
                        caseid_app.caseid_media33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media34':
                        caseid_app.caseid_media34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media35':
                        caseid_app.caseid_media35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media36':
                        caseid_app.caseid_media36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media37':
                        caseid_app.caseid_media37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media38':
                        caseid_app.caseid_media38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media39':
                        caseid_app.caseid_media39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media40':
                        caseid_app.caseid_media40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media41':
                        caseid_app.caseid_media41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media42':
                        caseid_app.caseid_media42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media43':
                        caseid_app.caseid_media43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media44':
                        caseid_app.caseid_media44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media45':
                        caseid_app.caseid_media45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media46':
                        caseid_app.caseid_media46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media47':
                        caseid_app.caseid_media47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media48':
                        caseid_app.caseid_media48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media49':
                        caseid_app.caseid_media49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media50':
                        caseid_app.caseid_media50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media51':
                        caseid_app.caseid_media51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media52':
                        caseid_app.caseid_media52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media53':
                        caseid_app.caseid_media53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media54':
                        caseid_app.caseid_media54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media55':
                        caseid_app.caseid_media55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media56':
                        caseid_app.caseid_media56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media57':
                        caseid_app.caseid_media57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media58':
                        caseid_app.caseid_media58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media59':
                        caseid_app.caseid_media59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media60':
                        caseid_app.caseid_media60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media61':
                        caseid_app.caseid_media61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media62':
                        caseid_app.caseid_media62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media63':
                        caseid_app.caseid_media63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media64':
                        caseid_app.caseid_media64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media65':
                        caseid_app.caseid_media65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media66':
                        caseid_app.caseid_media66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media67':
                        caseid_app.caseid_media67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Media68':
                        caseid_app.caseid_media68(self)
                except:
                    module_other_app.close_app()



#7 Báo cáo
def report(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 434
    while (rownum < 653):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Report01':
                        caseid_app.caseid_report01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report02':
                        caseid_app.caseid_report02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report03':
                        caseid_app.caseid_report03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report04':
                        caseid_app.caseid_report04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report05':
                        caseid_app.caseid_report05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report06':
                        caseid_app.caseid_report06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report07':
                        caseid_app.caseid_report07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report08':
                        caseid_app.caseid_report08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report09':
                        caseid_app.caseid_report09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report10':
                        caseid_app.caseid_report10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report11':
                        caseid_app.caseid_report11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report12':
                        caseid_app.caseid_report12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report13':
                        caseid_app.caseid_report13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report14':
                        caseid_app.caseid_report14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report15':
                        caseid_app.caseid_report15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report16':
                        caseid_app.caseid_report16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report17':
                        caseid_app.caseid_report17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report18':
                        caseid_app.caseid_report18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report19':
                        caseid_app.caseid_report19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report20':
                        caseid_app.caseid_report20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report21':
                        caseid_app.caseid_report21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report22':
                        caseid_app.caseid_report22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report23':
                        caseid_app.caseid_report23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report24':
                        caseid_app.caseid_report24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report25':
                        caseid_app.caseid_report25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report26':
                        caseid_app.caseid_report26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report27':
                        caseid_app.caseid_report27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report28':
                        caseid_app.caseid_report28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report29':
                        caseid_app.caseid_report29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report30':
                        caseid_app.caseid_report30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report31':
                        caseid_app.caseid_report31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report32':
                        caseid_app.caseid_report32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report33':
                        caseid_app.caseid_report33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report34':
                        caseid_app.caseid_report34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report35':
                        caseid_app.caseid_report35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report36':
                        caseid_app.caseid_report36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report37':
                        caseid_app.caseid_report37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report38':
                        caseid_app.caseid_report38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report39':
                        caseid_app.caseid_report39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report40':
                        caseid_app.caseid_report40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report41':
                        caseid_app.caseid_report41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report42':
                        caseid_app.caseid_report42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report43':
                        caseid_app.caseid_report43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report44':
                        caseid_app.caseid_report44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report45':
                        caseid_app.caseid_report45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report46':
                        caseid_app.caseid_report46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report47':
                        caseid_app.caseid_report47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report48':
                        caseid_app.caseid_report48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report49':
                        caseid_app.caseid_report49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report50':
                        caseid_app.caseid_report50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report51':
                        caseid_app.caseid_report51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report52':
                        caseid_app.caseid_report52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report53':
                        caseid_app.caseid_report53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report54':
                        caseid_app.caseid_report54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report55':
                        caseid_app.caseid_report55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report56':
                        caseid_app.caseid_report56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report57':
                        caseid_app.caseid_report57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report58':
                        caseid_app.caseid_report58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report59':
                        caseid_app.caseid_report59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report60':
                        caseid_app.caseid_report60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report61':
                        caseid_app.caseid_report61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report62':
                        caseid_app.caseid_report62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report63':
                        caseid_app.caseid_report63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report64':
                        caseid_app.caseid_report64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report65':
                        caseid_app.caseid_report65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report66':
                        caseid_app.caseid_report66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report67':
                        caseid_app.caseid_report67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report68':
                        caseid_app.caseid_report68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report69':
                        caseid_app.caseid_report69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report70':
                        caseid_app.caseid_report70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report71':
                        caseid_app.caseid_report71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report72':
                        caseid_app.caseid_report72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report73':
                        caseid_app.caseid_report73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report74':
                        caseid_app.caseid_report74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report75':
                        caseid_app.caseid_report75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report76':
                        caseid_app.caseid_report76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report77':
                        caseid_app.caseid_report77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report78':
                        caseid_app.caseid_report78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report79':
                        caseid_app.caseid_report79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report80':
                        caseid_app.caseid_report80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report81':
                        caseid_app.caseid_report81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report82':
                        caseid_app.caseid_report82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report83':
                        caseid_app.caseid_report83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report84':
                        caseid_app.caseid_report84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report85':
                        caseid_app.caseid_report85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report86':
                        caseid_app.caseid_report86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report87':
                        caseid_app.caseid_report87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report88':
                        caseid_app.caseid_report88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report89':
                        caseid_app.caseid_report89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report90':
                        caseid_app.caseid_report90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report91':
                        caseid_app.caseid_report91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report92':
                        caseid_app.caseid_report92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report93':
                        caseid_app.caseid_report93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report94':
                        caseid_app.caseid_report94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report95':
                        caseid_app.caseid_report95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report96':
                        caseid_app.caseid_report96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report97':
                        caseid_app.caseid_report97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report98':
                        caseid_app.caseid_report98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report99':
                        caseid_app.caseid_report99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report100':
                        caseid_app.caseid_report100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report101':
                        caseid_app.caseid_report101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report102':
                        caseid_app.caseid_report102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report103':
                        caseid_app.caseid_report103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report104':
                        caseid_app.caseid_report104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report105':
                        caseid_app.caseid_report105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report106':
                        caseid_app.caseid_report106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report107':
                        caseid_app.caseid_report107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report108':
                        caseid_app.caseid_report108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report109':
                        caseid_app.caseid_report109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report110':
                        caseid_app.caseid_report110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report111':
                        caseid_app.caseid_report111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report112':
                        caseid_app.caseid_report112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report113':
                        caseid_app.caseid_report113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report114':
                        caseid_app.caseid_report114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report115':
                        caseid_app.caseid_report115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report116':
                        caseid_app.caseid_report116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report117':
                        caseid_app.caseid_report117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report118':
                        caseid_app.caseid_report118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report119':
                        caseid_app.caseid_report119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report120':
                        caseid_app.caseid_report120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report121':
                        caseid_app.caseid_report121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report122':
                        caseid_app.caseid_report122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report123':
                        caseid_app.caseid_report123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report124':
                        caseid_app.caseid_report124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report125':
                        caseid_app.caseid_report125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report126':
                        caseid_app.caseid_report126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report127':
                        caseid_app.caseid_report127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report128':
                        caseid_app.caseid_report128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report129':
                        caseid_app.caseid_report129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report130':
                        caseid_app.caseid_report130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report131':
                        caseid_app.caseid_report131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report132':
                        caseid_app.caseid_report132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report133':
                        caseid_app.caseid_report133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report134':
                        caseid_app.caseid_report134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report135':
                        caseid_app.caseid_report135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report136':
                        caseid_app.caseid_report136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report137':
                        caseid_app.caseid_report137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report138':
                        caseid_app.caseid_report138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report139':
                        caseid_app.caseid_report139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report140':
                        caseid_app.caseid_report140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report141':
                        caseid_app.caseid_report141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report142':
                        caseid_app.caseid_report142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report143':
                        caseid_app.caseid_report143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report144':
                        caseid_app.caseid_report144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report145':
                        caseid_app.caseid_report145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report146':
                        caseid_app.caseid_report146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report147':
                        caseid_app.caseid_report147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report148':
                        caseid_app.caseid_report148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report149':
                        caseid_app.caseid_report149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report150':
                        caseid_app.caseid_report150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report151':
                        caseid_app.caseid_report151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report152':
                        caseid_app.caseid_report152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report153':
                        caseid_app.caseid_report153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report154':
                        caseid_app.caseid_report154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report155':
                        caseid_app.caseid_report155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report156':
                        caseid_app.caseid_report156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report157':
                        caseid_app.caseid_report157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report158':
                        caseid_app.caseid_report158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report159':
                        caseid_app.caseid_report159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report160':
                        caseid_app.caseid_report160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report161':
                        caseid_app.caseid_report161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report162':
                        caseid_app.caseid_report162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report163':
                        caseid_app.caseid_report163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report164':
                        caseid_app.caseid_report164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report165':
                        caseid_app.caseid_report165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report166':
                        caseid_app.caseid_report166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report167':
                        caseid_app.caseid_report167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report168':
                        caseid_app.caseid_report168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report169':
                        caseid_app.caseid_report169(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Report01':
                        caseid_app.caseid_report01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report02':
                        caseid_app.caseid_report02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report03':
                        caseid_app.caseid_report03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report04':
                        caseid_app.caseid_report04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report05':
                        caseid_app.caseid_report05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report06':
                        caseid_app.caseid_report06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report07':
                        caseid_app.caseid_report07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report08':
                        caseid_app.caseid_report08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report09':
                        caseid_app.caseid_report09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report10':
                        caseid_app.caseid_report10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report11':
                        caseid_app.caseid_report11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report12':
                        caseid_app.caseid_report12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report13':
                        caseid_app.caseid_report13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report14':
                        caseid_app.caseid_report14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report15':
                        caseid_app.caseid_report15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report16':
                        caseid_app.caseid_report16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report17':
                        caseid_app.caseid_report17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report18':
                        caseid_app.caseid_report18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report19':
                        caseid_app.caseid_report19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report20':
                        caseid_app.caseid_report20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report21':
                        caseid_app.caseid_report21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report22':
                        caseid_app.caseid_report22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report23':
                        caseid_app.caseid_report23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report24':
                        caseid_app.caseid_report24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report25':
                        caseid_app.caseid_report25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report26':
                        caseid_app.caseid_report26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report27':
                        caseid_app.caseid_report27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report28':
                        caseid_app.caseid_report28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report29':
                        caseid_app.caseid_report29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report30':
                        caseid_app.caseid_report30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report31':
                        caseid_app.caseid_report31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report32':
                        caseid_app.caseid_report32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report33':
                        caseid_app.caseid_report33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report34':
                        caseid_app.caseid_report34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report35':
                        caseid_app.caseid_report35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report36':
                        caseid_app.caseid_report36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report37':
                        caseid_app.caseid_report37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report38':
                        caseid_app.caseid_report38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report39':
                        caseid_app.caseid_report39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report40':
                        caseid_app.caseid_report40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report41':
                        caseid_app.caseid_report41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report42':
                        caseid_app.caseid_report42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report43':
                        caseid_app.caseid_report43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report44':
                        caseid_app.caseid_report44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report45':
                        caseid_app.caseid_report45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report46':
                        caseid_app.caseid_report46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report47':
                        caseid_app.caseid_report47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report48':
                        caseid_app.caseid_report48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report49':
                        caseid_app.caseid_report49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report50':
                        caseid_app.caseid_report50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report51':
                        caseid_app.caseid_report51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report52':
                        caseid_app.caseid_report52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report53':
                        caseid_app.caseid_report53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report54':
                        caseid_app.caseid_report54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report55':
                        caseid_app.caseid_report55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report56':
                        caseid_app.caseid_report56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report57':
                        caseid_app.caseid_report57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report58':
                        caseid_app.caseid_report58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report59':
                        caseid_app.caseid_report59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report60':
                        caseid_app.caseid_report60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report61':
                        caseid_app.caseid_report61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report62':
                        caseid_app.caseid_report62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report63':
                        caseid_app.caseid_report63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report64':
                        caseid_app.caseid_report64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report65':
                        caseid_app.caseid_report65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report66':
                        caseid_app.caseid_report66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report67':
                        caseid_app.caseid_report67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report68':
                        caseid_app.caseid_report68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report69':
                        caseid_app.caseid_report69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report70':
                        caseid_app.caseid_report70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report71':
                        caseid_app.caseid_report71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report72':
                        caseid_app.caseid_report72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report73':
                        caseid_app.caseid_report73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report74':
                        caseid_app.caseid_report74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report75':
                        caseid_app.caseid_report75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report76':
                        caseid_app.caseid_report76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report77':
                        caseid_app.caseid_report77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report78':
                        caseid_app.caseid_report78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report79':
                        caseid_app.caseid_report79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report80':
                        caseid_app.caseid_report80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report81':
                        caseid_app.caseid_report81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report82':
                        caseid_app.caseid_report82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report83':
                        caseid_app.caseid_report83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report84':
                        caseid_app.caseid_report84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report85':
                        caseid_app.caseid_report85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report86':
                        caseid_app.caseid_report86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report87':
                        caseid_app.caseid_report87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report88':
                        caseid_app.caseid_report88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report89':
                        caseid_app.caseid_report89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report90':
                        caseid_app.caseid_report90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report91':
                        caseid_app.caseid_report91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report92':
                        caseid_app.caseid_report92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report93':
                        caseid_app.caseid_report93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report94':
                        caseid_app.caseid_report94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report95':
                        caseid_app.caseid_report95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report96':
                        caseid_app.caseid_report96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report97':
                        caseid_app.caseid_report97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report98':
                        caseid_app.caseid_report98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report99':
                        caseid_app.caseid_report99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report100':
                        caseid_app.caseid_report100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report101':
                        caseid_app.caseid_report101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report102':
                        caseid_app.caseid_report102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report103':
                        caseid_app.caseid_report103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report104':
                        caseid_app.caseid_report104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report105':
                        caseid_app.caseid_report105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report106':
                        caseid_app.caseid_report106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report107':
                        caseid_app.caseid_report107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report108':
                        caseid_app.caseid_report108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report109':
                        caseid_app.caseid_report109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report110':
                        caseid_app.caseid_report110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report111':
                        caseid_app.caseid_report111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report112':
                        caseid_app.caseid_report112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report113':
                        caseid_app.caseid_report113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report114':
                        caseid_app.caseid_report114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report115':
                        caseid_app.caseid_report115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report116':
                        caseid_app.caseid_report116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report117':
                        caseid_app.caseid_report117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report118':
                        caseid_app.caseid_report118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report119':
                        caseid_app.caseid_report119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report120':
                        caseid_app.caseid_report120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report121':
                        caseid_app.caseid_report121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report122':
                        caseid_app.caseid_report122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report123':
                        caseid_app.caseid_report123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report124':
                        caseid_app.caseid_report124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report125':
                        caseid_app.caseid_report125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report126':
                        caseid_app.caseid_report126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report127':
                        caseid_app.caseid_report127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report128':
                        caseid_app.caseid_report128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report129':
                        caseid_app.caseid_report129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report130':
                        caseid_app.caseid_report130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report131':
                        caseid_app.caseid_report131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report132':
                        caseid_app.caseid_report132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report133':
                        caseid_app.caseid_report133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report134':
                        caseid_app.caseid_report134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report135':
                        caseid_app.caseid_report135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report136':
                        caseid_app.caseid_report136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report137':
                        caseid_app.caseid_report137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report138':
                        caseid_app.caseid_report138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report139':
                        caseid_app.caseid_report139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report140':
                        caseid_app.caseid_report140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report141':
                        caseid_app.caseid_report141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report142':
                        caseid_app.caseid_report142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report143':
                        caseid_app.caseid_report143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report144':
                        caseid_app.caseid_report144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report145':
                        caseid_app.caseid_report145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report146':
                        caseid_app.caseid_report146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report147':
                        caseid_app.caseid_report147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report148':
                        caseid_app.caseid_report148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report149':
                        caseid_app.caseid_report149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report150':
                        caseid_app.caseid_report150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report151':
                        caseid_app.caseid_report151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report152':
                        caseid_app.caseid_report152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report153':
                        caseid_app.caseid_report153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report154':
                        caseid_app.caseid_report154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report155':
                        caseid_app.caseid_report155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report156':
                        caseid_app.caseid_report156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report157':
                        caseid_app.caseid_report157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report158':
                        caseid_app.caseid_report158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report159':
                        caseid_app.caseid_report159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report160':
                        caseid_app.caseid_report160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report161':
                        caseid_app.caseid_report161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report162':
                        caseid_app.caseid_report162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report163':
                        caseid_app.caseid_report163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report164':
                        caseid_app.caseid_report164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report165':
                        caseid_app.caseid_report165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report166':
                        caseid_app.caseid_report166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report167':
                        caseid_app.caseid_report167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report168':
                        caseid_app.caseid_report168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report169':
                        caseid_app.caseid_report169(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Report01':
                        caseid_app.caseid_report01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report02':
                        caseid_app.caseid_report02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report03':
                        caseid_app.caseid_report03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report04':
                        caseid_app.caseid_report04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report05':
                        caseid_app.caseid_report05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report06':
                        caseid_app.caseid_report06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report07':
                        caseid_app.caseid_report07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report08':
                        caseid_app.caseid_report08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report09':
                        caseid_app.caseid_report09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report10':
                        caseid_app.caseid_report10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report11':
                        caseid_app.caseid_report11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report12':
                        caseid_app.caseid_report12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report13':
                        caseid_app.caseid_report13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report14':
                        caseid_app.caseid_report14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report15':
                        caseid_app.caseid_report15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report16':
                        caseid_app.caseid_report16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report17':
                        caseid_app.caseid_report17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report18':
                        caseid_app.caseid_report18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report19':
                        caseid_app.caseid_report19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report20':
                        caseid_app.caseid_report20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report21':
                        caseid_app.caseid_report21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report22':
                        caseid_app.caseid_report22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report23':
                        caseid_app.caseid_report23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report24':
                        caseid_app.caseid_report24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report25':
                        caseid_app.caseid_report25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report26':
                        caseid_app.caseid_report26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report27':
                        caseid_app.caseid_report27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report28':
                        caseid_app.caseid_report28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report29':
                        caseid_app.caseid_report29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report30':
                        caseid_app.caseid_report30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report31':
                        caseid_app.caseid_report31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report32':
                        caseid_app.caseid_report32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report33':
                        caseid_app.caseid_report33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report34':
                        caseid_app.caseid_report34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report35':
                        caseid_app.caseid_report35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report36':
                        caseid_app.caseid_report36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report37':
                        caseid_app.caseid_report37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report38':
                        caseid_app.caseid_report38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report39':
                        caseid_app.caseid_report39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report40':
                        caseid_app.caseid_report40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report41':
                        caseid_app.caseid_report41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report42':
                        caseid_app.caseid_report42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report43':
                        caseid_app.caseid_report43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report44':
                        caseid_app.caseid_report44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report45':
                        caseid_app.caseid_report45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report46':
                        caseid_app.caseid_report46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report47':
                        caseid_app.caseid_report47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report48':
                        caseid_app.caseid_report48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report49':
                        caseid_app.caseid_report49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report50':
                        caseid_app.caseid_report50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report51':
                        caseid_app.caseid_report51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report52':
                        caseid_app.caseid_report52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report53':
                        caseid_app.caseid_report53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report54':
                        caseid_app.caseid_report54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report55':
                        caseid_app.caseid_report55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report56':
                        caseid_app.caseid_report56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report57':
                        caseid_app.caseid_report57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report58':
                        caseid_app.caseid_report58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report59':
                        caseid_app.caseid_report59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report60':
                        caseid_app.caseid_report60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report61':
                        caseid_app.caseid_report61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report62':
                        caseid_app.caseid_report62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report63':
                        caseid_app.caseid_report63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report64':
                        caseid_app.caseid_report64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report65':
                        caseid_app.caseid_report65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report66':
                        caseid_app.caseid_report66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report67':
                        caseid_app.caseid_report67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report68':
                        caseid_app.caseid_report68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report69':
                        caseid_app.caseid_report69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report70':
                        caseid_app.caseid_report70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report71':
                        caseid_app.caseid_report71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report72':
                        caseid_app.caseid_report72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report73':
                        caseid_app.caseid_report73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report74':
                        caseid_app.caseid_report74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report75':
                        caseid_app.caseid_report75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report76':
                        caseid_app.caseid_report76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report77':
                        caseid_app.caseid_report77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report78':
                        caseid_app.caseid_report78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report79':
                        caseid_app.caseid_report79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report80':
                        caseid_app.caseid_report80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report81':
                        caseid_app.caseid_report81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report82':
                        caseid_app.caseid_report82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report83':
                        caseid_app.caseid_report83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report84':
                        caseid_app.caseid_report84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report85':
                        caseid_app.caseid_report85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report86':
                        caseid_app.caseid_report86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report87':
                        caseid_app.caseid_report87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report88':
                        caseid_app.caseid_report88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report89':
                        caseid_app.caseid_report89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report90':
                        caseid_app.caseid_report90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report91':
                        caseid_app.caseid_report91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report92':
                        caseid_app.caseid_report92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report93':
                        caseid_app.caseid_report93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report94':
                        caseid_app.caseid_report94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report95':
                        caseid_app.caseid_report95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report96':
                        caseid_app.caseid_report96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report97':
                        caseid_app.caseid_report97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report98':
                        caseid_app.caseid_report98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report99':
                        caseid_app.caseid_report99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report100':
                        caseid_app.caseid_report100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report101':
                        caseid_app.caseid_report101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report102':
                        caseid_app.caseid_report102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report103':
                        caseid_app.caseid_report103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report104':
                        caseid_app.caseid_report104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report105':
                        caseid_app.caseid_report105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report106':
                        caseid_app.caseid_report106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report107':
                        caseid_app.caseid_report107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report108':
                        caseid_app.caseid_report108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report109':
                        caseid_app.caseid_report109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report110':
                        caseid_app.caseid_report110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report111':
                        caseid_app.caseid_report111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report112':
                        caseid_app.caseid_report112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report113':
                        caseid_app.caseid_report113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report114':
                        caseid_app.caseid_report114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report115':
                        caseid_app.caseid_report115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report116':
                        caseid_app.caseid_report116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report117':
                        caseid_app.caseid_report117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report118':
                        caseid_app.caseid_report118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report119':
                        caseid_app.caseid_report119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report120':
                        caseid_app.caseid_report120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report121':
                        caseid_app.caseid_report121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report122':
                        caseid_app.caseid_report122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report123':
                        caseid_app.caseid_report123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report124':
                        caseid_app.caseid_report124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report125':
                        caseid_app.caseid_report125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report126':
                        caseid_app.caseid_report126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report127':
                        caseid_app.caseid_report127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report128':
                        caseid_app.caseid_report128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report129':
                        caseid_app.caseid_report129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report130':
                        caseid_app.caseid_report130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report131':
                        caseid_app.caseid_report131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report132':
                        caseid_app.caseid_report132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report133':
                        caseid_app.caseid_report133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report134':
                        caseid_app.caseid_report134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report135':
                        caseid_app.caseid_report135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report136':
                        caseid_app.caseid_report136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report137':
                        caseid_app.caseid_report137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report138':
                        caseid_app.caseid_report138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report139':
                        caseid_app.caseid_report139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report140':
                        caseid_app.caseid_report140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report141':
                        caseid_app.caseid_report141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report142':
                        caseid_app.caseid_report142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report143':
                        caseid_app.caseid_report143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report144':
                        caseid_app.caseid_report144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report145':
                        caseid_app.caseid_report145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report146':
                        caseid_app.caseid_report146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report147':
                        caseid_app.caseid_report147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report148':
                        caseid_app.caseid_report148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report149':
                        caseid_app.caseid_report149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report150':
                        caseid_app.caseid_report150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report151':
                        caseid_app.caseid_report151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report152':
                        caseid_app.caseid_report152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report153':
                        caseid_app.caseid_report153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report154':
                        caseid_app.caseid_report154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report155':
                        caseid_app.caseid_report155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report156':
                        caseid_app.caseid_report156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report157':
                        caseid_app.caseid_report157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report158':
                        caseid_app.caseid_report158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report159':
                        caseid_app.caseid_report159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report160':
                        caseid_app.caseid_report160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report161':
                        caseid_app.caseid_report161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report162':
                        caseid_app.caseid_report162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report163':
                        caseid_app.caseid_report163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report164':
                        caseid_app.caseid_report164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report165':
                        caseid_app.caseid_report165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report166':
                        caseid_app.caseid_report166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report167':
                        caseid_app.caseid_report167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report168':
                        caseid_app.caseid_report168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report169':
                        caseid_app.caseid_report169(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Report01':
                        caseid_app.caseid_report01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report02':
                        caseid_app.caseid_report02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report03':
                        caseid_app.caseid_report03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report04':
                        caseid_app.caseid_report04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report05':
                        caseid_app.caseid_report05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report06':
                        caseid_app.caseid_report06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report07':
                        caseid_app.caseid_report07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report08':
                        caseid_app.caseid_report08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report09':
                        caseid_app.caseid_report09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report10':
                        caseid_app.caseid_report10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report11':
                        caseid_app.caseid_report11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report12':
                        caseid_app.caseid_report12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report13':
                        caseid_app.caseid_report13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report14':
                        caseid_app.caseid_report14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report15':
                        caseid_app.caseid_report15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report16':
                        caseid_app.caseid_report16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report17':
                        caseid_app.caseid_report17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report18':
                        caseid_app.caseid_report18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report19':
                        caseid_app.caseid_report19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report20':
                        caseid_app.caseid_report20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report21':
                        caseid_app.caseid_report21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report22':
                        caseid_app.caseid_report22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report23':
                        caseid_app.caseid_report23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report24':
                        caseid_app.caseid_report24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report25':
                        caseid_app.caseid_report25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report26':
                        caseid_app.caseid_report26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report27':
                        caseid_app.caseid_report27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report28':
                        caseid_app.caseid_report28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report29':
                        caseid_app.caseid_report29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report30':
                        caseid_app.caseid_report30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report31':
                        caseid_app.caseid_report31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report32':
                        caseid_app.caseid_report32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report33':
                        caseid_app.caseid_report33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report34':
                        caseid_app.caseid_report34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report35':
                        caseid_app.caseid_report35(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report36':
                        caseid_app.caseid_report36(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report37':
                        caseid_app.caseid_report37(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report38':
                        caseid_app.caseid_report38(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report39':
                        caseid_app.caseid_report39(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report40':
                        caseid_app.caseid_report40(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report41':
                        caseid_app.caseid_report41(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report42':
                        caseid_app.caseid_report42(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report43':
                        caseid_app.caseid_report43(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report44':
                        caseid_app.caseid_report44(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report45':
                        caseid_app.caseid_report45(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report46':
                        caseid_app.caseid_report46(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report47':
                        caseid_app.caseid_report47(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report48':
                        caseid_app.caseid_report48(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report49':
                        caseid_app.caseid_report49(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report50':
                        caseid_app.caseid_report50(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report51':
                        caseid_app.caseid_report51(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report52':
                        caseid_app.caseid_report52(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report53':
                        caseid_app.caseid_report53(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report54':
                        caseid_app.caseid_report54(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report55':
                        caseid_app.caseid_report55(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report56':
                        caseid_app.caseid_report56(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report57':
                        caseid_app.caseid_report57(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report58':
                        caseid_app.caseid_report58(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report59':
                        caseid_app.caseid_report59(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report60':
                        caseid_app.caseid_report60(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report61':
                        caseid_app.caseid_report61(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report62':
                        caseid_app.caseid_report62(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report63':
                        caseid_app.caseid_report63(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report64':
                        caseid_app.caseid_report64(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report65':
                        caseid_app.caseid_report65(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report66':
                        caseid_app.caseid_report66(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report67':
                        caseid_app.caseid_report67(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report68':
                        caseid_app.caseid_report68(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report69':
                        caseid_app.caseid_report69(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report70':
                        caseid_app.caseid_report70(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report71':
                        caseid_app.caseid_report71(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report72':
                        caseid_app.caseid_report72(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report73':
                        caseid_app.caseid_report73(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report74':
                        caseid_app.caseid_report74(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report75':
                        caseid_app.caseid_report75(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report76':
                        caseid_app.caseid_report76(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report77':
                        caseid_app.caseid_report77(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report78':
                        caseid_app.caseid_report78(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report79':
                        caseid_app.caseid_report79(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report80':
                        caseid_app.caseid_report80(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report81':
                        caseid_app.caseid_report81(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report82':
                        caseid_app.caseid_report82(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report83':
                        caseid_app.caseid_report83(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report84':
                        caseid_app.caseid_report84(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report85':
                        caseid_app.caseid_report85(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report86':
                        caseid_app.caseid_report86(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report87':
                        caseid_app.caseid_report87(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report88':
                        caseid_app.caseid_report88(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report89':
                        caseid_app.caseid_report89(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report90':
                        caseid_app.caseid_report90(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report91':
                        caseid_app.caseid_report91(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report92':
                        caseid_app.caseid_report92(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report93':
                        caseid_app.caseid_report93(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report94':
                        caseid_app.caseid_report94(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report95':
                        caseid_app.caseid_report95(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report96':
                        caseid_app.caseid_report96(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report97':
                        caseid_app.caseid_report97(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report98':
                        caseid_app.caseid_report98(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report99':
                        caseid_app.caseid_report99(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report100':
                        caseid_app.caseid_report100(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report101':
                        caseid_app.caseid_report101(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report102':
                        caseid_app.caseid_report102(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report103':
                        caseid_app.caseid_report103(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report104':
                        caseid_app.caseid_report104(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report105':
                        caseid_app.caseid_report105(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report106':
                        caseid_app.caseid_report106(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report107':
                        caseid_app.caseid_report107(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report108':
                        caseid_app.caseid_report108(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report109':
                        caseid_app.caseid_report109(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report110':
                        caseid_app.caseid_report110(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report111':
                        caseid_app.caseid_report111(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report112':
                        caseid_app.caseid_report112(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report113':
                        caseid_app.caseid_report113(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report114':
                        caseid_app.caseid_report114(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report115':
                        caseid_app.caseid_report115(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report116':
                        caseid_app.caseid_report116(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report117':
                        caseid_app.caseid_report117(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report118':
                        caseid_app.caseid_report118(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report119':
                        caseid_app.caseid_report119(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report120':
                        caseid_app.caseid_report120(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report121':
                        caseid_app.caseid_report121(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report122':
                        caseid_app.caseid_report122(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report123':
                        caseid_app.caseid_report123(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report124':
                        caseid_app.caseid_report124(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report125':
                        caseid_app.caseid_report125(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report126':
                        caseid_app.caseid_report126(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report127':
                        caseid_app.caseid_report127(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report128':
                        caseid_app.caseid_report128(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report129':
                        caseid_app.caseid_report129(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report130':
                        caseid_app.caseid_report130(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report131':
                        caseid_app.caseid_report131(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report132':
                        caseid_app.caseid_report132(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report133':
                        caseid_app.caseid_report133(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report134':
                        caseid_app.caseid_report134(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report135':
                        caseid_app.caseid_report135(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report136':
                        caseid_app.caseid_report136(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report137':
                        caseid_app.caseid_report137(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report138':
                        caseid_app.caseid_report138(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report139':
                        caseid_app.caseid_report139(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report140':
                        caseid_app.caseid_report140(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report141':
                        caseid_app.caseid_report141(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report142':
                        caseid_app.caseid_report142(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report143':
                        caseid_app.caseid_report143(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report144':
                        caseid_app.caseid_report144(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report145':
                        caseid_app.caseid_report145(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report146':
                        caseid_app.caseid_report146(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report147':
                        caseid_app.caseid_report147(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report148':
                        caseid_app.caseid_report148(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report149':
                        caseid_app.caseid_report149(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report150':
                        caseid_app.caseid_report150(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report151':
                        caseid_app.caseid_report151(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report152':
                        caseid_app.caseid_report152(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report153':
                        caseid_app.caseid_report153(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report154':
                        caseid_app.caseid_report154(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report155':
                        caseid_app.caseid_report155(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report156':
                        caseid_app.caseid_report156(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report157':
                        caseid_app.caseid_report157(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report158':
                        caseid_app.caseid_report158(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report159':
                        caseid_app.caseid_report159(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report160':
                        caseid_app.caseid_report160(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report161':
                        caseid_app.caseid_report161(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report162':
                        caseid_app.caseid_report162(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report163':
                        caseid_app.caseid_report163(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report164':
                        caseid_app.caseid_report164(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report165':
                        caseid_app.caseid_report165(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report166':
                        caseid_app.caseid_report166(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report167':
                        caseid_app.caseid_report167(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report168':
                        caseid_app.caseid_report168(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Report169':
                        caseid_app.caseid_report169(self)
                except:
                    module_other_app.close_app()



#8 Tiện ích
def utilitie(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_app.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 654
    while (rownum < 708):
        rownum += 1
        rownum = str(rownum)
        if sheet["D"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["E"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["F"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["G"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo2)

    modetest = ''.join(re.findall(r'\d+', var_app.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Utilities01':
                        caseid_app.caseid_utilities01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities02':
                        caseid_app.caseid_utilities02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities03':
                        caseid_app.caseid_utilities03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities04':
                        caseid_app.caseid_utilities04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities05':
                        caseid_app.caseid_utilities05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities06':
                        caseid_app.caseid_utilities06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities07':
                        caseid_app.caseid_utilities07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities08':
                        caseid_app.caseid_utilities08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities09':
                        caseid_app.caseid_utilities09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities10':
                        caseid_app.caseid_utilities10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities11':
                        caseid_app.caseid_utilities11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities12':
                        caseid_app.caseid_utilities12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities13':
                        caseid_app.caseid_utilities13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities14':
                        caseid_app.caseid_utilities14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities15':
                        caseid_app.caseid_utilities15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities16':
                        caseid_app.caseid_utilities16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities17':
                        caseid_app.caseid_utilities17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities18':
                        caseid_app.caseid_utilities18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities19':
                        caseid_app.caseid_utilities19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities20':
                        caseid_app.caseid_utilities20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities21':
                        caseid_app.caseid_utilities21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities22':
                        caseid_app.caseid_utilities22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities23':
                        caseid_app.caseid_utilities23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities24':
                        caseid_app.caseid_utilities24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities25':
                        caseid_app.caseid_utilities25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities26':
                        caseid_app.caseid_utilities26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities27':
                        caseid_app.caseid_utilities27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities28':
                        caseid_app.caseid_utilities28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities29':
                        caseid_app.caseid_utilities29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities30':
                        caseid_app.caseid_utilities30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities31':
                        caseid_app.caseid_utilities31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities32':
                        caseid_app.caseid_utilities32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities33':
                        caseid_app.caseid_utilities33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities34':
                        caseid_app.caseid_utilities34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities35':
                        caseid_app.caseid_utilities35(self)
                except:
                    module_other_app.close_app()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Utilities01':
                        caseid_app.caseid_utilities01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities02':
                        caseid_app.caseid_utilities02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities03':
                        caseid_app.caseid_utilities03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities04':
                        caseid_app.caseid_utilities04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities05':
                        caseid_app.caseid_utilities05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities06':
                        caseid_app.caseid_utilities06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities07':
                        caseid_app.caseid_utilities07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities08':
                        caseid_app.caseid_utilities08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities09':
                        caseid_app.caseid_utilities09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities10':
                        caseid_app.caseid_utilities10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities11':
                        caseid_app.caseid_utilities11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities12':
                        caseid_app.caseid_utilities12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities13':
                        caseid_app.caseid_utilities13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities14':
                        caseid_app.caseid_utilities14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities15':
                        caseid_app.caseid_utilities15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities16':
                        caseid_app.caseid_utilities16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities17':
                        caseid_app.caseid_utilities17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities18':
                        caseid_app.caseid_utilities18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities19':
                        caseid_app.caseid_utilities19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities20':
                        caseid_app.caseid_utilities20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities21':
                        caseid_app.caseid_utilities21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities22':
                        caseid_app.caseid_utilities22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities23':
                        caseid_app.caseid_utilities23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities24':
                        caseid_app.caseid_utilities24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities25':
                        caseid_app.caseid_utilities25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities26':
                        caseid_app.caseid_utilities26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities27':
                        caseid_app.caseid_utilities27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities28':
                        caseid_app.caseid_utilities28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities29':
                        caseid_app.caseid_utilities29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities30':
                        caseid_app.caseid_utilities30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities31':
                        caseid_app.caseid_utilities31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities32':
                        caseid_app.caseid_utilities32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities33':
                        caseid_app.caseid_utilities33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities34':
                        caseid_app.caseid_utilities34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities35':
                        caseid_app.caseid_utilities35(self)
                except:
                    module_other_app.close_app()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Utilities01':
                        caseid_app.caseid_utilities01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities02':
                        caseid_app.caseid_utilities02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities03':
                        caseid_app.caseid_utilities03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities04':
                        caseid_app.caseid_utilities04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities05':
                        caseid_app.caseid_utilities05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities06':
                        caseid_app.caseid_utilities06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities07':
                        caseid_app.caseid_utilities07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities08':
                        caseid_app.caseid_utilities08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities09':
                        caseid_app.caseid_utilities09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities10':
                        caseid_app.caseid_utilities10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities11':
                        caseid_app.caseid_utilities11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities12':
                        caseid_app.caseid_utilities12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities13':
                        caseid_app.caseid_utilities13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities14':
                        caseid_app.caseid_utilities14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities15':
                        caseid_app.caseid_utilities15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities16':
                        caseid_app.caseid_utilities16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities17':
                        caseid_app.caseid_utilities17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities18':
                        caseid_app.caseid_utilities18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities19':
                        caseid_app.caseid_utilities19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities20':
                        caseid_app.caseid_utilities20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities21':
                        caseid_app.caseid_utilities21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities22':
                        caseid_app.caseid_utilities22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities23':
                        caseid_app.caseid_utilities23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities24':
                        caseid_app.caseid_utilities24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities25':
                        caseid_app.caseid_utilities25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities26':
                        caseid_app.caseid_utilities26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities27':
                        caseid_app.caseid_utilities27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities28':
                        caseid_app.caseid_utilities28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities29':
                        caseid_app.caseid_utilities29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities30':
                        caseid_app.caseid_utilities30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities31':
                        caseid_app.caseid_utilities31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities32':
                        caseid_app.caseid_utilities32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities33':
                        caseid_app.caseid_utilities33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities34':
                        caseid_app.caseid_utilities34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities35':
                        caseid_app.caseid_utilities35(self)
                except:
                    module_other_app.close_app()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Utilities01':
                        caseid_app.caseid_utilities01(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities02':
                        caseid_app.caseid_utilities02(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities03':
                        caseid_app.caseid_utilities03(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities04':
                        caseid_app.caseid_utilities04(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities05':
                        caseid_app.caseid_utilities05(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities06':
                        caseid_app.caseid_utilities06(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities07':
                        caseid_app.caseid_utilities07(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities08':
                        caseid_app.caseid_utilities08(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities09':
                        caseid_app.caseid_utilities09(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities10':
                        caseid_app.caseid_utilities10(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities11':
                        caseid_app.caseid_utilities11(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities12':
                        caseid_app.caseid_utilities12(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities13':
                        caseid_app.caseid_utilities13(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities14':
                        caseid_app.caseid_utilities14(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities15':
                        caseid_app.caseid_utilities15(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities16':
                        caseid_app.caseid_utilities16(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities17':
                        caseid_app.caseid_utilities17(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities18':
                        caseid_app.caseid_utilities18(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities19':
                        caseid_app.caseid_utilities19(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities20':
                        caseid_app.caseid_utilities20(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities21':
                        caseid_app.caseid_utilities21(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities22':
                        caseid_app.caseid_utilities22(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities23':
                        caseid_app.caseid_utilities23(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities24':
                        caseid_app.caseid_utilities24(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities25':
                        caseid_app.caseid_utilities25(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities26':
                        caseid_app.caseid_utilities26(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities27':
                        caseid_app.caseid_utilities27(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities28':
                        caseid_app.caseid_utilities28(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities29':
                        caseid_app.caseid_utilities29(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities30':
                        caseid_app.caseid_utilities30(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities31':
                        caseid_app.caseid_utilities31(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities32':
                        caseid_app.caseid_utilities32(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities33':
                        caseid_app.caseid_utilities33(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities34':
                        caseid_app.caseid_utilities34(self)
                except:
                    module_other_app.close_app()
                try:
                    if case == 'Utilities35':
                        caseid_app.caseid_utilities35(self)
                except:
                    module_other_app.close_app()


